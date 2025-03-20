import pandas as pd
import os, sys, warnings, re
import pdr.handlers.Console_Handler as console
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell
from oracledb import Connection
from colorama import Fore, Style

# Description: This is a utility module that contains all the utility functions that are used in the Type 4 Report job.


def check_null_empty(value) -> bool:
    """
    Check if a value is None, empty, or contains no elements.

    Args:
        value (str, list, tuple): A value which can be a string, list, or tuple

    Returns:
        bool: True if the value is None, an empty string, or an empty collection, False otherwise
    """
    # Check if the value is None
    if value is None:
        return True
    # Check for string and use strip to remove any whitespace
    if isinstance(value, str):
        return value.strip() == ""
    # Check for list or tuple and if they are empty
    if isinstance(value, (list, tuple)):
        return len(value) == 0
    return False


def check_list_length(list: list, check_length: int):
    """Check if the list is null or empty and if the length of the list is equal to the specified length.

    Args:
        list (list): The list to check.
        check_length (int): The specified length to check against.

    Raises:
        ValueError: If the list is None or empty.
        ValueError: If the length of the list is less than the specified length.
    """
    # Check if the list is null or empty
    if check_null_empty(list):
        raise ValueError("The list is None or empty.")
    # Check if the length of the list is equal to the specified length
    if len(list) < check_length:
        raise ValueError(f"The list length is less than {check_length}.")
    if len(list) > check_length:
        console.log(f"The list length is greater than {check_length}.")


def _clean_text(text: str):
    """
    Clean the text by removing illegal characters (control chars except \t, \n, \r).
    """
    # Remove illegal characters
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)


def dataframe_to_excel(
    df: pd.DataFrame,
    ws: Worksheet,
    skip_rows: int = 0,
    skip_cols: int = 0,
    sample_cell: Cell = None,
    alt_border: bool = False,
    clean_text: bool = False,
):
    """Copy the data from a DataFrame to a worksheet starting from the specified cell.

    Args:
        df (pd.DataFrame): The DataFrame to be copied to the worksheet
        ws (Worksheet): The worksheet to copy the DataFrame to
        skip_rows (int, optional): _description_. Defaults to 0.
        skip_cols (int, optional): _description_. Defaults to 0.
        sample_cell (Cell, optional): _description_. Defaults to None.
        alt_border (bool, optional): _description_. Defaults to False.
        clean_text (bool, optional): _description_. Defaults to False.
    """
    max_col = len(df.columns)
    max_row = len(df.index)
    for col in range(1, max_col + 1):
        for row in range(1, max_row + 1):
            col_complete = get_column_letter(col + skip_cols)
            row_complete = row + skip_rows
            dest_cell = ws[col_complete + str(row_complete)]
            # Copy the style of the sample cell to the destination cell if it is provided
            if sample_cell is not None:
                copy_paste_cell(
                    sample_cell, dest_cell, alt_value=True, alt_border=alt_border
                )
            df_value = df.iat[row - 1, col - 1]
            if clean_text:
                data_type = str(type(df_value)).lower()
                df_value = _clean_text(str(df_value))
                if "int" in data_type:
                    df_value = int(df_value)
                elif "float" in data_type:
                    df_value = float(df_value)
                elif "timestamp" in data_type:
                    df_value = pd.Timestamp(df_value)
                elif "none" in data_type:
                    df_value = ""
            dest_cell.value = df_value
            

def copy_paste_cell(
    src_cell: Cell,
    dest_cell: Cell,
    alt_value: bool = False,
    alt_font: bool = False,
    alt_alignment: bool = False,
    alt_fill: bool = False,
    alt_border: bool = False,
):
    """Copy and pasted the style and value of a cell to another cell.

    Args:
        src_cell (Cell): The cell to copy from.
        dest_cell (Cell): The cell to paste to.
        alt_value (bool, optional): False if not provide alternative value to set up. Defaults to False.
        alt_font (bool, optional): . Defaults to False.
        alt_alignment (bool, optional): _description_. Defaults to False.
        alt_fill (bool, optional): _description_. Defaults to False.
        alt_border (bool, optional): _description_. Defaults to False.
    """
    # Copy value and style of src_cell to dest_cell
    if not alt_value:
        dest_cell.value = src_cell.value
    if not alt_font:
        dest_cell.font = get_cell_font(src_cell)
    if not alt_alignment:
        dest_cell.alignment = get_cell_alignment(src_cell)
    if not alt_fill:
        dest_cell.fill = get_cell_fill(src_cell)
    if not alt_border:
        dest_cell.border = get_cell_border(src_cell)


def get_cell_fill(cell: Cell) -> PatternFill:
    """Get the fill of a cell.

    Args:
        cell (Cell): The cell to get the fill from.

    Returns:
        PatternFill: The fill of the cell.
    """
    if cell.fill.fgColor.type == "rgb":
        fill = PatternFill(
            start_color=cell.fill.start_color,
            end_color=cell.fill.end_color,
            fill_type=cell.fill.fill_type,
        )
    else:
        console.log(f"Cell {cell.coordinate} does not have a fill color.")
    return fill


def get_cell_border(cell: Cell) -> Border:
    """Get the border of a cell.

    Args:
        cell (Cell): The cell to get the border from.

    Returns:
        Border: The border of the cell.
    """
    return Border(
        left=Side(
            border_style=cell.border.left.border_style, color=cell.border.left.color
        ),
        right=Side(
            border_style=cell.border.right.border_style,
            color=cell.border.right.color,
        ),
        top=Side(
            border_style=cell.border.top.border_style, color=cell.border.top.color
        ),
        bottom=Side(
            border_style=cell.border.bottom.border_style,
            color=cell.border.bottom.color,
        ),
    )


def get_cell_font(cell: Cell) -> Font:
    """Get the font of a cell.

    Args:
        cell (Cell): The cell to get the font from.

    Returns:
        Font: The font of the cell.
    """
    return Font(
        name=cell.font.name,
        size=cell.font.size,
        bold=cell.font.bold,
        italic=cell.font.italic,
        vertAlign=cell.font.vertAlign,
        underline=cell.font.underline,
        strike=cell.font.strike,
        color=cell.font.color,
    )


def get_cell_alignment(cell: Cell) -> Alignment:
    """Get the alignment of a cell.

    Args:
        cell (Cell): The cell to get the alignment from.

    Returns:
        Alignment: The alignment of the cell.
    """
    return Alignment(
        horizontal=cell.alignment.horizontal,
        vertical=cell.alignment.vertical,
        text_rotation=cell.alignment.text_rotation,
        wrap_text=cell.alignment.wrap_text,
        shrink_to_fit=cell.alignment.shrink_to_fit,
        indent=cell.alignment.indent,
    )
    

def set_cell_style(
    cell,
    data,
    name="Arial",
    size=10,
    bold=False,
    underline=None,
    vertical="bottom",
    horizontal="center",
    wrapText=False,
    fill=None,
    border=None,
):
    try: 
        if data is not None:
            cell.value = data
        cell.font = Font(name=name, size=size, bold=bold, underline=underline)
        cell.alignment = Alignment(
            vertical=vertical, horizontal=horizontal, wrapText=wrapText
        )
        if fill is not None:
            cell.fill = fill
        if border is not None:
            cell.border = border
    except Exception as e:
        console.log(Fore.RED + f"Error in set_cell_style(): {e}" + Style.RESET_ALL)
        raise e
    

def close_wb(wb: Workbook, output_file_name: str):
    """Close and save the workbook with the formatted output file name

    Args:
        wb (Workbook): the workbook object
        output_file_name (str): the formatted output file name
    """
    try:
        # Save the workbook
        wb.save(f"{output_file_name}")
        wb.close()
    except Exception as e:
        console.log(Fore.RED + f"Error when saving and closing the workbook: {e}" + Style.RESET_ALL)
        raise e
    

def auto_adjust_column_width(ws: Worksheet, start_row: int = 1):
    """Auto adjust the column width of the worksheet based on the longest string of each column, starting from a specified row.

    Args:
        ws (Worksheet): The worksheet to adjust the column width.
        start_row (int): The first row to start scanning from (1-based index).
    """
    column_widths = {}
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value and str(cell.value)[0] != "=":
                # Update the width of the column if this cell's content is wider
                column_width = len(str(cell.value))
                if column_width > column_widths.get(cell.column_letter, 0):
                    if cell.row == 1:
                        column_width += 4
                    column_widths[cell.column_letter] = column_width

            if cell.row >= 100000:
                # Stop scanning if the row is greater than 100000
                raise ValueError("The row is greater than 100000, Stopping scanning. Please check template file for Phantom data.")

    # Adding a small buffer to column width for aesthetics
    buffer = 2
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width + buffer


def delete_files(list_delete: list):
    """Delete the files in the list.

    Args:
        list_delete (list): The list of files to delete.
    """
    try:
        # Iterate through the list of files to delete
        for file_path in list_delete:
            # Check if file exist or not
            if os.path.exists(file_path):
                # Remove the specified file path
                os.remove(file_path)
                console.log(f"File '{file_path}' has been deleted successfully.")
            else:
                console.log(f"File '{file_path}' does not exist.")
    except FileNotFoundError:
        console.log(f"File {file_path} does not exist.")
    except PermissionError:
        console.log(f"Permission denied: unable to delete '{file_path}'.")
    except Exception as e:
        console.log(f"Error deleting the file {file_path}: {e}")


def read_sql_file(file_path: str) -> str:
    """Read the SQL file and return the content as a string.

    Args:
        file_path (str): The file path of the SQL file.

    Returns:
        str: The content of the SQL file as a string.
    """
    with open(file_path, "r") as file:
        sql = file.read()
    # Remove any leading and trailing whitespaces
    sql = sql.strip()
    return sql
        

def get_client_X_quarter(connection: Connection, period_code: int) -> int:
    """
    Obtain client_X defined quarter from TXXX2P. Takes in Oracle connection, period code.
    """
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore")

            query = f"select QTR_PERIOD_TITLE from pm_dm_dom.dlvr_time where PERIOD_CODE = {period_code}"
            qtr_df = pd.read_sql_query(query, connection)
            current_qtr = qtr_df["QTR_PERIOD_TITLE"].iloc[0]
            current_qtr = int(current_qtr[4])
        return current_qtr
    except Exception as e:
        console.log("client_X quarter error: " + str(e))
        sys.exit(1)


def get_client_X_quarter_code(connection: Connection) -> int:
    """
    Obtain most recent client_X defined quarter rollover period code from TXXX2P. Takes in Oracle connection.
    """
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore")
            query = """ with cpc as (
                            select max(period_code) pc from pm_dm_dom.dlvr_time
                        ), cqc as (
                            select quarter_code qc from pm_dm_dom.dlvr_time where period_code = (select pc from cpc)
                        )
                        select min(period_code) QCODE from pm_dm_dom.dlvr_time where quarter_code = (select qc from cqc) """
            qtr_df = pd.read_sql_query(query, connection)
            quarter_code = qtr_df.iat[0, 0]
        return quarter_code
    except Exception as e:
        console.log("client_X quarter code error: " + str(e))
        sys.exit(1)
        
        
def check_file_exist(file_path_list: list):
    """Check if the files in the list exist.

    Args:
        file_path_list (list): The list of file paths to check.

    Raises:
        FileNotFoundError: If the file is not found.
    """
    if check_null_empty(file_path_list):
        raise ValueError(Fore.RED + "The file path list is None or empty." + Style.RESET_ALL)
    for path in file_path_list:
        if not os.path.exists(path):
            raise FileNotFoundError(Fore.RED + f"File not found: {path}" + Style.RESET_ALL)
        

def get_df_from_sql(connection: Connection, sql_path: str, query: str) -> pd.DataFrame:
    """Execute the SQL query in the file and return the result as a pandas DataFrame.

    Args:
        connection (Connection): The Oracle connection object.
        sql_path (str): The file path of the SQL query.

    Raises:
        ValueError: If the query in the file is empty.

    Returns:
        pd.DataFrame: The result of the SQL query as a pandas DataFrame.
    """
    # Check if the query is empty
    if check_null_empty(query):
        raise ValueError(f"Query in {sql_path} is empty.")
    # If the last character is a ;, remove it
    if query.endswith(";"):
        query = query[:-1]
    # Execute the query and get the result into a pandas dataframe
    df = pd.read_sql(query, connection)
    #print(df)
    console.log(f"Retrieved Dataframe by executing the query in {sql_path}")
    # Check if the dataframe is empty
    if df is None or df.empty:
        console.log(Fore.YELLOW + f"WARNING: Dataframe is empty as a result of executing sql queries in {sql_path}" + Style.RESET_ALL)
    return df


def execute_oracle_package(procedure: str, args: list, connection: Connection):
    """Execute a stored procedure or function in the Oracle database

    Args:
        procedure (str): The name of the stored procedure or function
        args (list): The arguments to pass to the stored procedure or function
        connection (Connection): The Oracle connection object
    """
    with connection.cursor() as cursor:
        cursor.callproc(procedure, args)
    connection.commit()