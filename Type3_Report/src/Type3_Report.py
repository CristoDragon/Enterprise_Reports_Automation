import pandas as pd
import pdr.period.Altria as alt_prd
import pdr.handlers.Console_Handler as console
import warnings
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell


# Author: Dragon Xu
# Date: 06/04/2024
# Description: This class is used to pull and process data from the Oracle database for Type3_Rpt project.
# It contains necessary methods to create Excel worksheets reports.


class Type3_Report:
    # Add slots to reduce memory usage and improve attribute access speed
    __slots__ = (
        'connection', 'schema_usr', 'schema_usr_prev', 'table_time', 'table_brand',
        'table_store', 'list_ws_names', 'list_category_code', 'list_category_desc',
        'dict_category', 'dict_a1', 'skip_rows_field_sheet', 'skip_rows_drill_sheet',
        'skip_rows_main_sheet', 'skip_cols_main_sheet', 'skip_rows_cat_home_sheet',
        'skip_rows_main_adjusted', 'yellow_fill', 'gray_fill', 'thin_border',
        'thick_border', 'white_font', 'period', 'end_week'
    )
    
    def __init__(self, connection):
        # Initialize the connection to the Oracle database
        self.connection = connection
        # Initialize the schema names and table names used in the queries
        self.schema_usr = "PM_DM_DOM"
        self.schema_usr_prev = "PM_DM_DOM_PREV"
        self.table_time = "dlvr_time"
        self.table_brand = "dlvr_brand"
        self.table_store = "dlvr_store"

        # Initialize worksheet names in the template file
        self.list_ws_names = ["Main_temp", "Cat_temp", "Drill_temp", "Fld_temp"]
        # Initialize category codes and names
        self.list_category_code = [3123, 3292, 3211, 3217, 3225, 3293, 3262]
        self.list_category_desc = [
            ["Cigarettes", "Cig"],
            ["E-Cigarettes", "ECig"],
            ["Moist", "MST"],
            ["SNUS", "SNUS"],
            ["Cigars", "CGR"],
            ["TDP", "TDP"],
            ["Tobacco Accessories", "Accessories"],
        ]
        self.dict_category = dict(zip(self.list_category_code, self.list_category_desc))

        # Initialize dictionaries for each of the category above
        # Note that each of the dictionary represents all the column-category mapping, to protect the confidential information, only the first column is shown here.
        a1_cig = {
            "CATEGORY": ["category_desc", "Category"],   
        }
        a1_ecig = {
            "ECIG_LEVEL_CODE": ["ecig_level_desc", "Nicotine Mg"],    
        }
        a1_moist = {
            "PRICE_TIER": ["price_tier_desc", "Price Tier"],
        }
        a1_snus = {
            "CATEGORY": ["category_desc", "Category"],
        }
        a1_cigars = {
            "CATEGORY": ["category_desc", "Category"],
        }
        a1_tdp = {
            "CATEGORY": ["category_desc", "Category"],
        }
        a1_access = {
            "CATEGORY": ["category_desc", "Category"],
        }
        # Initialize a dictionary to link category codes to attribute dictionaries
        self.dict_a1 = {
            3123: a1_cig,
            3292: a1_ecig,
            3211: a1_moist,
            3217: a1_snus,
            3225: a1_cigars,
            3293: a1_tdp,
            3262: a1_access,
        }

        # Initialize skip rows constant
        self.skip_rows_field_sheet = 8
        self.skip_rows_drill_sheet = 5
        self.skip_rows_main_sheet = 8
        self.skip_cols_main_sheet = 2
        self.skip_rows_cat_home_sheet = 8
        self.skip_rows_main_adjusted = 8

        # Initialize the fill style
        self.yellow_fill = PatternFill(
            start_color="EEECE1", end_color="EEECE1", fill_type="solid"
        )
        self.gray_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        # Initialize border styles
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")
        self.thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)
        # Initialize the font style
        self.white_font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")

        # Initialize the period code and end week
        self.period, self.end_week = self.get_period_info()
        console.log(f"Period: {self.period}, End Week: {self.end_week}")

        # Suppress the sqlalchemy database connection warning
        warnings.filterwarnings(
            "ignore", message="pandas only supports SQLAlchemy connectable"
        )

    @staticmethod
    def check_null_empty(value) -> bool:
        """
        Check if a value is None, empty, or contains no elements.

        Args:
            value (str, list, tuple): A value which can be a string, list, or tuple

        Returns:
            bool: True if the value is None, an empty string, or an empty collection, False otherwise
        """
        # Check if the value is None
        if value is None:
            return True

        # Check for string and use strip to remove any whitespace
        if isinstance(value, str):
            return value.strip() == ""

        # Check for list or tuple and if they are empty
        if isinstance(value, (list, tuple)):
            return len(value) == 0

        return False

    def get_period_info(self) -> tuple:
        """Return the period code and end week for the current period.

        Returns:
            tuple: A tuple containing the period code and end week
        """
        try:
            period = alt_prd.get_altria_period_code(self.connection)
            end_week = alt_prd.get_altria_end_week(self.connection, period)
            return (period, end_week)
        except Exception as e:
            console.log(
                f"Failed to retrieve period code or end_week from the Oracle database: {e}"
            )
            raise e

    def pull_attributes(self, c: str, c1: int, a1: str, a2: str) -> tuple:
        """Pull and process data for the specified category and attributes.

        Args:
            c (str): the category code surrounded by % signs
            c1 (int): the category code
            a1 (str): the attribute name
            a2 (str): the attribute name with '_desc' appended

        Returns:
            tuple: a tuple of two DataFrames, one for the merged data and one for the summary data
        """
        # Initialize the names of count columns
        list_c_p = ["c", "p"]
        # Initialize new column names for flags
        list_new_cols = ["NEW", "DROP", "SAME", "DESC_CHG"]
        list_reorder = ["DROP", "DESC_CHG", "NEW", "SAME"]
        
        try:
            # Execute queries to pull raw data into two DataFrames, one for current period and one for previous period
            df_current = self.pull_raw_data(
                f"{self.schema_usr}.{self.table_brand}", a1, a2, c, list_c_p[0]
            )
            df_previous = self.pull_raw_data(
                f"{self.schema_usr_prev}.{self.table_brand}", a1, a2, c, list_c_p[1]
            )

            # Merge the two dataframes above, created new columns for flags, and return the merged DataFrame
            df_merged = self.merge_df(df_previous, df_current, a1.upper())
            df_merged = self.process_merged_df(df_merged, a2.upper(), c1, list_new_cols)
            # Calculate the summary info for the merged DataFrame
            df_summary = self.cal_sum(df_merged, list_reorder)
            # Collapse each attribute to get summary info and return the dataframe
            return (df_merged, df_summary)
        except Exception as e:
            console.log(f"Failed in pull_attributes() for {c1} {a1}: {e}")
            raise e

    def pull_raw_data(
        self, table: str, a1: str, a2: str, c: str, c_p: str
    ) -> pd.DataFrame:
        """
        Pull raw data from the specified table for the given columns, filter for category.

        Args:
            table (str): The name of the table to pull data from
            a1 (str): The name of the first columns
            a2 (str): The name of the second columns
            c (str): The category code
            c_p (str): The prefix indicating current or previous period

        Returns:
            pd.DataFrame: A DataFrame containing the raw data from the specified table
        """
        try: 
            # Validate input
            Type3_Report._validate_pull_raw_data(
                table, a1, a2, c, c_p, "pull_raw_data"
            )
            # Define the SQL queries
            query = f"""
            SELECT {a1}, {a2} AS {c_p}{a2}, COUNT(*) AS {c_p}cnt
            FROM {table}
            WHERE source_1 = '1' AND category LIKE '{c}'
            GROUP BY {a1}, {a2}
            ORDER BY {a1}
            """
            # Execute queries, fetch data into DataFrame, and return the DataFrame
            return pd.read_sql_query(query, self.connection)
        except Exception as e:
            console.log(f"Failed to execute query: {query}. Error: {e}")
            raise e

    @staticmethod
    def _validate_pull_raw_data(
        table: pd.DataFrame,
        a1: str,
        a2: str,
        c: str,
        c_p: str,
        method_name: str = "pull_raw_data",
    ):
        try:
            Type3_Report.validate_str_list_tuple(table, method_name)
            Type3_Report.validate_str_list_tuple(a1, method_name)
            Type3_Report.validate_str_list_tuple(a2, method_name)
            Type3_Report.validate_str_list_tuple(c, method_name)
            Type3_Report.validate_str_list_tuple(c_p, method_name)
        except Exception as e:
            console.log(f"Error in {method_name}: {e}")
            raise e

    def merge_df(
        self, df_previous: pd.DataFrame, df_current: pd.DataFrame, a1: str
    ) -> pd.DataFrame:
        """Merge the two dataframes on the specified column and return the merged dataframe.

        Args:
            df_previous (pd.DataFrame): the dataframe for the previous period
            df_current (pd.DataFrame): the dataframe for the current period
            a1 (str): the column to merge on

        Raises:
            ValueError: if the input data frames are None or empty

        Returns:
            pd.DataFrame: the merged dataframe
        """
        try:
            # Validate input data frames
            Type3_Report._validate_input_merge_df(df_previous, df_current, a1)

            # Merge the data frames on column a1 and return it
            return pd.merge(
                df_previous,
                df_current,
                on=a1,
                how="outer",
                suffixes=("_PREV", "_CURR"),  # note that the prefix is not used
            )
        except Exception as e:
            console.log(f"Failed to merge dataframes for {a1}: {e}")
            raise e

    @staticmethod
    def _validate_input_merge_df(
        df_previous: pd.DataFrame,
        df_current: pd.DataFrame,
        a1: str,
        method_name: str = "merge_df",
    ):
        try:
            Type3_Report.validate_df(df_previous, method_name)
            Type3_Report.validate_df(df_current, method_name)
            Type3_Report.validate_str_list_tuple(a1, method_name)
        except Exception as e:
            console.log(f"Error in {method_name}(): {e}")
            raise e

    def process_merged_df(
        self, df_merged: pd.DataFrame, a2: str, c1: int, list_new_cols: list
    ) -> pd.DataFrame:
        """Process the merged dataframe to add new columns for flags and return the processed dataframe.

        Args:
            df_merged (pd.DataFrame): the merged dataframe to be processed
            a2 (str): the column to calculate the difference
            c1 (int): the category code
            list_new_cols (list): the list of new column names for flags

        Returns:
            pd.DataFrame: the processed dataframe with the new columns
        """
        # Initialize columns for ccnt and pcnt
        ccnt = "CCNT"
        pcnt = "PCNT"
        # Replace NaN with 0 for columns ["CCNT", "PCNT"] in order to execute numeric calculations
        df_merged[ccnt] = df_merged[ccnt].fillna(0)
        df_merged[pcnt] = df_merged[pcnt].fillna(0)
        # Create column diff
        df_merged["DIFF"] = df_merged[ccnt] - df_merged[pcnt]
        # Create column cat for category
        df_merged["CAT"] = c1
        # Create 4 columns for flags: ["DROP", "DESC_CHG", "NEW", "SAME"]
        df_merged[list_new_cols[0]] = (
            (df_merged[ccnt] > 0) & (df_merged[pcnt] == 0)
        ).astype(int)
        df_merged[list_new_cols[1]] = (
            (df_merged[ccnt] == 0) & (df_merged[pcnt] > 0)
        ).astype(int)
        df_merged[list_new_cols[2]] = (
            (df_merged[ccnt] > 0)
            & (df_merged[pcnt] > 0)
            & (df_merged[f"C{a2}"] == df_merged[f"P{a2}"])
        ).astype(int)
        df_merged[list_new_cols[3]] = (
            (df_merged[ccnt] > 0)
            & (df_merged[pcnt] > 0)
            & (df_merged[f"C{a2}"] != df_merged[f"P{a2}"])
        ).astype(int)
        return df_merged

    def cal_sum(self, df: pd.DataFrame, list_new_cols: list) -> pd.DataFrame:
        try:
            # Group by the specified category and calculate sum for the flags
            summary_df = (
                df.groupby("CAT")
                .agg({col: "sum" for col in list_new_cols})
                .reset_index()
            )
        except KeyError as e:
            console.log(f"Column or key error: {e}")
            raise KeyError(f"Column or key {e} not found in DataFrame") from e
        except ValueError as e:
            console.log(f"Value error: {e}")
            raise ValueError("Incorrect value used in operation") from e
        except IndexError as e:
            console.log(f"Index error: {e}")
            raise IndexError("Accessed index does not exist") from e
        # Output results can be adjusted here, like exporting to a file or further transformations
        return summary_df

    def pull_drill_down_data(self, c: str, dict_cols: dict) -> pd.DataFrame:
        """Pull drill down data for the specified category and columns.

        Args:
            c (str): The category code
            dict_cols (dict): A dictionary containing the columns to be selected

        Returns:
            pd.DataFrame: A DataFrame containing the drill down data
        """
        try:
            # Validate input
            self._validate_pull_drill_down(c, dict_cols)

            # Initialize the columns to be selected
            # Ensure each key-value pair is concatenated with a comma and separated from the next pair
            list_cols = ["msa_brand_code, brand_title"]
            list_cols.extend(f"{k}, {v[0]}" for k, v in dict_cols.items())
        except KeyError as e:
            console.log(f"Key error: {e}")
            raise KeyError(f"Key {e} not found in dictionary") from e
        except Exception as e:
            console.log(f"Error in pull_drill_down_data(): {e}")
            raise e
        # Add additional columns for specific category codes
        if c == "%3123%":
            list_cols.extend(["color_family_code", "color_family_desc"])
        # Join all columns with commas to form the complete column string for the SQL query
        cols = ", ".join(list_cols)
        # Initialize the SQL query
        query = f"""SELECT DISTINCT {cols} 
                    FROM {self.schema_usr}.{self.table_brand}
                    WHERE category LIKE '{c}' AND source_1 = '1'"""
        # Execute queries, fetch data into DataFrame, and return the DataFrame
        try:
            return pd.read_sql_query(query, self.connection)
        except Exception as e:
            console.log(f"Failed to execute query: {query}. Error: {e}")
            raise RuntimeError(f"Failed to execute query due to an error: {e}") from e

    def _validate_pull_drill_down(
        self, c: str, dict_cols: dict, method_name="pull_drill_down_data"
    ):
        Type3_Report.validate_str_list_tuple(c, method_name)

        if dict_cols is None or len(dict_cols) == 0:
            raise ValueError(
                f"Invalid input for {method_name}(): dict_cols is None or empty"
            )

    @staticmethod
    def dataframe_to_excel(
        df: pd.DataFrame,
        ws: Worksheet,
        skip_rows: int = 0,
        skip_cols: int = 0,
        sample_cell: Cell = None,
        alt_border: bool = False,
    ):
        """Copy the data from a DataFrame to a worksheet starting from the specified cell.

        Args:
            df (pd.DataFrame): The DataFrame to be copied to the worksheet
            ws (Worksheet): The worksheet to copy the DataFrame to
            skip_rows (int, optional): _description_. Defaults to 0.
            skip_cols (int, optional): _description_. Defaults to 0.
            sample_cell (Cell, optional): _description_. Defaults to None.
            alt_border (bool, optional): _description_. Defaults to False.
        """
        max_col = len(df.columns)
        max_row = len(df.index)
        for col in range(1, max_col + 1):
            for row in range(1, max_row + 1):
                col_complete = get_column_letter(col + skip_cols)
                row_complete = row + skip_rows
                dest_cell = ws[col_complete + str(row_complete)]
                # Copy the style of the sample cell to the destination cell if it is provided
                if sample_cell is not None:
                    Type3_Report.copy_paste_cell(
                        sample_cell, dest_cell, alt_value=True, alt_border=alt_border
                    )
                dest_cell.value = df.iat[row - 1, col - 1]

    def create_main_sheet(self, wb: Workbook, sheet_name: str) -> Worksheet:
        """Create the main sheet with the specified name and return the worksheet.
        Args:
            wb (Workbook): The workbook object
            sheet_name (str): The name of the main sheet

        Returns:
            Worksheet: The updated main sheet
        """
        # Copy the template main home sheet to a new worksheet
        ws_main = wb.copy_worksheet(wb[self.list_ws_names[0]])
        # Rename the copied worksheet
        ws_main.title = sheet_name
        # Format the end week to be in the format "Month Day, Year"
        end_week = self.end_week.strftime("%B %d, %Y")
        # Update the end week in cell A2
        ws_main["A2"].value = f"Week Ending {end_week}"
        # Return the main sheet
        return ws_main

    def create_field_sheet(
        self,
        df: pd.DataFrame,
        wb: Workbook,
        sheet_name: str,
        c1: int,
        category_name: str,
        a1: str,
    ):
        """Create a field sheet with the specified name and add data and style.

        Args:
            df (pd.DataFrame): The DataFrame to be copied to the worksheet
            wb (Workbook): The workbook object
            sheet_name (str): The name of the field sheet
            c1 (int): The category code
            category_name (str): The name of the category
            a1 (str): The attribute name
        """
        try:
            # Validate input
            Type3_Report._validate_field_sheet_input(
                df, wb, sheet_name, category_name
            )

            # Copy the template field sheet to a new worksheet
            ws = wb.copy_worksheet(wb[self.list_ws_names[3]])
            # Rename the copied worksheet to the specified sheet name
            ws.title = sheet_name

            # Replace all the zeros in column ''
            # Place the cleaned dataframe into the worksheet starting from cell A9
            Type3_Report.dataframe_to_excel(
                df, ws, skip_rows=self.skip_rows_field_sheet
            )
            # Set the style for the worksheet
            self._set_style_field_sheet(ws, category_name, c1, a1)
        except Exception as e:
            console.log(f"Failed in create_field_sheet: {e}")
            raise e
        else:
            console.log(f"'{sheet_name}' has been created successfully")

    @staticmethod
    def _validate_field_sheet_input(
        df: pd.DataFrame,
        wb: Workbook,
        sheet_name: str,
        category_name: str,
        method_name: str = "create_field_sheet",
    ):
        """Validate the input for creating the field sheet.

        Args:
            df (pd.DataFrame): The DataFrame to be copied to the worksheet
            wb (Workbook): The workbook object
            sheet_name (str): The name of the field sheet
            category_name (str): The name of the category
            method_name (str, optional): _description_. Defaults to "create_field_sheet".
        """
        Type3_Report.validate_df(df, method_name)
        Type3_Report.validate_wb(wb, method_name)
        Type3_Report.validate_str_list_tuple(sheet_name, method_name)
        Type3_Report.validate_str_list_tuple(category_name, method_name)

    def _set_style_field_sheet(
        self, ws: Worksheet, category_name: str, c1: int, a1: str
    ):
        """Set the style for the field sheet.

        Args:
            ws (Worksheet): The worksheet to set the style for
            category_name (str): The name of the category
            c1 (int): The category code
            a1 (str): The attribute name
        """
        try:
            # Write the sheet title and end week in cell A1 and A2
            self.write_endweek_title(ws, category_name, f"({str(c1)})")

            # Update the attribute name in cell A6
            ws["A6"].value = a1.capitalize()
            # Update the period code in B7, D7
            ws["B7"].value = f"Week {self.period - 1}"
            ws["D7"].value = f"Week {self.period}"
            # Set filter for the header row
            ws.auto_filter.ref = f"A{self.skip_rows_field_sheet}:F{ws.max_row}"
            # Replace all the zeros in column 'pcnt', 'ccnt'
            self._replace_zeros(ws)
            # Set the alignment for all the data cells
            self._set_alignment_field_sheet(ws)
        except Exception as e:
            console.log(f"Failed in _set_style_field_sheet(): {e}")
            raise e
        
    def _replace_zeros(self, ws: Worksheet):
        """Replace all the zeros in column 'pcnt', 'ccnt' with None.

        Args:
            ws (Worksheet): The worksheet to replace the zeros in
        """
        for row in range(self.skip_rows_field_sheet + 1, ws.max_row + 1):
            for col in (3, 5):
                cell = ws.cell(row, col)
                if cell.value == 0:
                    cell.value = None

    def _set_alignment_field_sheet(self, ws: Worksheet):
        """Set the alignment for the field sheet.

        Args:
            ws (Worksheet): The worksheet to set the alignment for
        """
        # Initialize the fills and fonts based on sample cells D2, D3, D4
        red_fill = Type3_Report.get_cell_fill(ws["D2"])
        font1 = Type3_Report.get_cell_font(ws["D2"])
        green_fill = Type3_Report.get_cell_fill(ws["D4"])
        font2 = Type3_Report.get_cell_font(ws["D4"])
        yellow_fill = Type3_Report.get_cell_fill(ws["D3"])
        font3 = Type3_Report.get_cell_font(ws["D3"])
        

        # Set alignment for column A, C, E, F to be center & bottom, starting form row 9
        for row in range(self.skip_rows_field_sheet + 1, ws.max_row + 1):
            cell_col2 = ws.cell(row, 2)
            cell_col3 = ws.cell(row, 3)
            cell_col4 = ws.cell(row, 4)
            cell_col5 = ws.cell(row, 5)
            cell_col2_fill = None
            cell_col2_font = None
            cell_col2_horizontal = None
            cell_col4_fill = None
            cell_col4_font = None
            cell_col4_horizontal = None
            
            if cell_col2.value != cell_col4.value:
                cell_col4_horizontal = "center"
                cell_col2_horizontal = "center"
                if cell_col3.value is None:
                    cell_col4_fill = green_fill
                    cell_col4_font = font2 
                elif cell_col5.value is None:
                    cell_col2_fill = red_fill
                    cell_col2_font = font1
                else:
                    cell_col2_fill = yellow_fill
                    cell_col2_font = font3
                    cell_col4_fill = yellow_fill
                    cell_col4_font = font3
            Type3_Report._set_cell_style(
                    cell=cell_col2,
                    data=None,
                    horizontal=cell_col2_horizontal,
                    fill=cell_col2_fill,
                    border=self.thin_border,
            )
            cell_col2.font = cell_col2_font
            Type3_Report._set_cell_style(
                    cell=cell_col4,
                    data=None,
                    horizontal=cell_col4_horizontal,
                    fill=cell_col4_fill,
                    border=self.thin_border,
            )
            cell_col4.font = cell_col4_font
            for col in (1, 3, 5, 6):
                cell = ws.cell(row, col)
                Type3_Report._set_cell_style(
                        cell=cell,
                        data=None,
                        horizontal="center",
                        border=self.thin_border,
                )

    @staticmethod
    def _set_cell_style(
        cell,
        data,
        name="Arial",
        size=10,
        bold=False,
        underline=None,
        vertical="bottom",
        horizontal="center",
        wrapText=False,
        fill=None,
        border=None,
        *args,
    ):
        try: 
            if data is not None:
                cell.value = data
            cell.font = Font(name=name, size=size, bold=bold, underline=underline)
            cell.alignment = Alignment(
                vertical=vertical, horizontal=horizontal, wrapText=wrapText
            )
            if fill is not None:
                cell.fill = fill
            if border is not None:
                cell.border = border
        except Exception as e:
            console.log(f"Error in _set_cell_style(): {e}")
            raise e

    def create_drill_sheet(
        self,
        df: pd.DataFrame,
        wb: Workbook,
        sheet_name: str,
        category_name: str,
        c1: int,
    ):
        """Create a drill sheet with the specified name and add data and style.

        Args:
            df (pd.DataFrame): The dataframe to be copied to the worksheet
            wb (Workbook): The workbook object that contains the drill sheet
            sheet_name (str): The specified name of the drill sheet
            category_name (str): The name of the category
            c1 (int): The category code
        """
        try:
            # Validate input
            Type3_Report._validate_drill_sheet_input(
                df, wb, sheet_name, category_name
            )

            # Copy the template drill sheet to a new worksheet
            ws = wb.copy_worksheet(wb[self.list_ws_names[2]])
            # Rename the copied worksheet to the specified sheet name
            ws.title = sheet_name

            # Place the cleaned dataframe into the worksheet starting from cell A6
            Type3_Report.dataframe_to_excel(df, ws, self.skip_rows_drill_sheet)
            # Set the font for the worksheet
            self._set_style_drill_sheet(ws, category_name, df.columns[2:], c1, df.shape)
        except Exception as e:
            console.log(f"Failed in create_drill_sheet(): {e}")
            raise e
        else:
            console.log(f"'{sheet_name}' has been created successfully.")

    @staticmethod
    def _validate_drill_sheet_input(
        df: pd.DataFrame,
        wb: Workbook,
        sheet_name: str,
        category_name: str,
        method_name: str = "create_drill_sheet",
    ):
        """Validate the input for create_drill_sheet() method.

        Args:
            df (pd.DataFrame): _description_
            wb (Workbook): _description_
            sheet_name (str): _description_
            category_name (str): _description_
            method_name (str, optional): _description_. Defaults to "create_drill_sheet".
        """
        Type3_Report.validate_df(df, method_name)
        Type3_Report.validate_wb(wb, method_name)
        Type3_Report.validate_str_list_tuple(sheet_name, method_name)
        Type3_Report.validate_str_list_tuple(category_name, method_name)

    def _set_style_drill_sheet(
        self,
        ws: Worksheet,
        category_name: str,
        column_names: list,
        c1: int,
        shape: list,
    ):
        """Set the style for the drill sheet.

        Args:
            ws (Worksheet): The drill sheet to set the style for
            category_name (str): The name of the category
            column_names (list): The list of column names
            c1 (int): The category code
            shape (list): The shape of the DataFrame
        """
        # Write the sheet title and end week in cell A1 and A2
        self.write_endweek_title(ws, category_name, "Attribute Drill Report")

        # Starting from row 4 and initial column index for 'C' (the 3rd column)
        start_row = 4
        start_col = 3
        # Iterate through the column names skipping the first two columns ['MSA_BRAND_CODE', 'BRAND_TITLE'] because they are fixed for all categories
        for i in range(0, len(column_names), 2):
            # Calculate the column letters for the current index
            col_index = start_col + i
            start_col_letter = ws.cell(row=1, column=col_index).column_letter
            end_col_letter = ws.cell(row=1, column=col_index + 1).column_letter
            # Construct the cell range to be merged
            cell_range = f"{start_col_letter}{start_row}:{end_col_letter}{start_row}"
            # Merge two cells into one dynamically based on the number of column names, starting from cell C4
            ws.merge_cells(cell_range)
            # Set the value and style for the merged cell in start_row
            for k, v in self.dict_a1[c1].items():
                if k == column_names[i]:
                    # Set the columns fill to be yellow for those who have a even column number, otherwise use default fill
                    fill1 = None
                    if i % 4 == 0:
                        fill1 = self.yellow_fill
                    # Set the value and style for the cell in start_row + 1
                    Type3_Report._set_cell_style(
                        cell=ws[cell_range.split(":")[0]],
                        data=v[1],
                        name="Courier New",
                        size=11,
                        bold=True,
                        vertical="center",
                        wrapText=True,
                        fill=fill1,
                        border=self.thick_border,
                    )
                    ws[cell_range.split(":")[1]].border = self.thick_border

            # Set the value and style for cell in start_row + 1 by copying the cell C5 and D5
            Type3_Report.copy_paste_cell(
                ws["C5"], ws.cell(row=5, column=col_index)
            )
            Type3_Report.copy_paste_cell(
                ws["D5"], ws.cell(row=5, column=col_index + 1)
            )

        # Set the style for all the data cells
        self._set_data_style_drill_sheet(ws, shape)

        # Set the column name for the last merged cell if the sheet is for category 3123
        if c1 == self.list_category_code[0]:
            Type3_Report._set_cell_style(
                cell=ws["AW4"],
                data="Color Family",
                name="Courier New",
                size=11,
                bold=True,
                vertical="center",
                wrapText=True,
                fill=fill1,
                border=self.thick_border,
            )
            ws["AX4"].border = self.thick_border

        # Set the filter for the header row in row 5
        col_filter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"C{self.skip_rows_drill_sheet}:{col_filter}{ws.max_row}"

    def _set_data_style_drill_sheet(self, ws: Worksheet, shape: list):
        """Set the style for the data cells in the drill sheet.

        Args:
            ws (Worksheet): The drill sheet to set the style for
            shape (list): The shape of the DataFrame
        """
        for row in range(6, shape[0] + 6):
            for col in range(1, 3):
                Type3_Report._set_cell_style(
                    cell=ws.cell(row=row, column=col),
                    data=None,
                    name="Courier New",
                    horizontal=None,
                    border=self.thin_border,
                )
            for col in range(3, shape[1] + 1):
                horizontal = "center"
                if col % 2 == 0:
                    horizontal = None
                Type3_Report._set_cell_style(
                    cell=ws.cell(row=row, column=col),
                    data=None,
                    name="Courier New",
                    horizontal=horizontal,
                    fill=self.yellow_fill,
                    border=self.thin_border,
                )

    def write_endweek_title(self, ws: Worksheet, category_name: str, report_name: str):
        """Write the end week and category title to the worksheet.

        Args:
            ws (Worksheet): The worksheet to write the title to.
            category_name (str): The category name
            report_name (str): The report name
        """
        # Format the end week to be in the format "Month Day, Year"
        end_week = self.end_week.strftime("%B %d, %Y")
        # Update the category title for the worksheet in cell A1 (only changing the value and keeping the style)
        ws["A1"].value = f"{category_name} {report_name}"
        # Update the end week in cell A2
        ws["A2"].value = f"Week Ending {end_week}"

    def add_data_main_sheet(self, df: pd.DataFrame, ws: Worksheet, c1: int):
        """Add the data from the DataFrame to the main sheet based on the category code.

        Args:
            df (pd.DataFrame): The summary data to be added to the main sheet.
            ws (Worksheet): The main sheet to add the data to.
            c1 (int): The category code to determine where to add the data.

        Raises:
            ValueError: If the category code is not found in the dictionary.
        """
        # Validate input
        self._validate_main_sheet_input(df, ws, c1, "add_data_main_sheet")

        # Find the index of c1 in the list of category codes
        try:
            index = self.list_category_code.index(c1)
        except ValueError:
            raise ValueError(
                f"Invalid input for add_data_main_sheet(): Category code {c1} is not found in list_category_code."
            )

        # Calculate skip_cols and skip_rows based on the index of c1
        skip_cols_adjusted = self.skip_cols_main_sheet + (index * 4)

        # Put the dataframe in excel if the cell has a white background color.
        # Otherwise, skip this cell until the next white cell and put the dataframe in the next white cell
        # Find the next white cell in the row
        while not Type3_Report.is_fill_default(
            ws.cell(row=self.skip_rows_main_adjusted + 1, column=skip_cols_adjusted + 1)
        ):
            self.skip_rows_main_adjusted += 1
        # Place the cleaned dataframe into the worksheet
        Type3_Report.dataframe_to_excel(
            df,
            ws,
            skip_rows=self.skip_rows_main_adjusted,
            skip_cols=skip_cols_adjusted,
            sample_cell=ws["C9"],
            alt_border=True,
        )
        # Iterate through the four cells in which the data were just added, color
        for col in range(skip_cols_adjusted + 1, skip_cols_adjusted + 4):
            # data_cell is the cell that has data added and whose to be checked
            data_cell = ws.cell(row=self.skip_rows_main_adjusted + 1, column=col)
            # header_cell is the in row 8 and in the same column as data_cell, indicates [DROP, DESC_CHG, NEW]
            header_cell = ws.cell(row=8, column=col)
            # Color the data_cell according to its header cell if the data is greater than 0
            if data_cell.value > 0:
                if header_cell.value == "Drop":
                    # Set the fill color according to cell Q2 in the main home sheet
                    data_cell.fill = Type3_Report.get_cell_fill(ws["Q2"])
                    data_cell.font = self.white_font
                elif header_cell.value == "Desc. Change":
                    # Set the fill color according to cell Q3 in the main home sheet
                    data_cell.fill = Type3_Report.get_cell_fill(ws["Q3"])
                elif header_cell.value == "New":
                    # Set the fill color according to cell Q4 in the main home sheet
                    data_cell.fill = Type3_Report.get_cell_fill(ws["Q4"])
                    data_cell.font = self.white_font

    def _validate_main_sheet_input(
        self,
        df: pd.DataFrame,
        ws: Worksheet,
        c1: int,
        method_name: str = "add_data_main_sheet",
    ):
        """Validate the input for the add_data_main_sheet() method.

        Args:
            df (pd.DataFrame): The DataFrame to be validated.
            ws (Worksheet): The worksheet to be validated.
            c1 (int): The category code to be validated.
            method_name (str, optional): The method where validation happened. Defaults to "add_data_main_sheet".

        Raises:
            ValueError: _description_
        """
        Type3_Report.validate_df(df, method_name)
        Type3_Report.validate_ws(ws, method_name, check_title="Main Home")
        if c1 not in self.dict_category.keys():
            raise ValueError(
                f"Invalid input for {method_name}(): Category code {c1} is not found in the dictionary."
            )

    @staticmethod
    def is_fill_default(cell: Cell):
        """
        Check if the background color of the given cell is not white.
        Assumes color is given in ARGB format (Alpha, Red, Green, Blue).

        Args:
            cell (openpyxl.cell.cell.Cell): The cell whose background color to check.

        Returns:
            bool: True if the cell has a default fill (typically white), False otherwise.
        """
        # Check if the fill is not set or is default
        if cell.fill.fill_type is None or cell.fill.fill_type == "none":
            # The cell is using the default settings, which is typically white
            return True
        return False

    def create_category_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        ws_main: Worksheet,
        c1: int,
        category_name: str,
    ):
        """Create a category home sheet for the specified category code and name.

        Args:
            wb (Workbook): The workbook object where the new sheet will be created.
            sheet_name (str): The name of the new sheet to be created.
            ws_main (Worksheet): The main home worksheet to copy data from.
            c1 (int): The category code.
            category_name (str): The category name.
        """
        try:
            # Validate input
            self._validate_category_sheet_input(wb, sheet_name)

            # Copy the template category home sheet to a new worksheet
            ws = wb.copy_worksheet(wb[self.list_ws_names[1]])
            # Rename the copied worksheet to the specified sheet name
            ws.title = sheet_name

            # Ad-hoc fix for category sheet title consistency
            if c1 == 3262:
                category_name = "Accessories"
            # Write the sheet title and end week in cell A1 and A2
            self.write_endweek_title(ws, category_name, "Summary")
            # Add data to columns G, H, I, and J by copying the data from the main home sheet
            self._put_data_GHIJ(ws, ws_main, c1)
            # Fill in all the "Not Applicable" rows with "N/A" and color them gray
            self._put_data_NA(ws)
            # Add data to columns: B, C, D from the instance dictionary variable
            self._put_data_BCD(ws, c1)
        except Exception as e:
            console.log(f"Failed in create_category_sheet(): {e}")
            raise e
        else:
            console.log(f"'{sheet_name}' has been created successfully.")

    def _put_data_GHIJ(self, ws: Worksheet, ws_main: Worksheet, c1: int):
        """Add data to columns G, H, I, and J by copying the data from the main home sheet.

        Args:
            ws (Worksheet): The worksheet to add data to.
            ws_main (Worksheet): The main home worksheet to copy data from.
            c1 (int): The category code.
        """
        # Initialize the dictionary of attribute names and codes for the specified category code
        dest_range = "G8:J33"
        cell_g10 = ws["G10"]
        # Add data to columns G, H, I, and J by copying the data from the main home sheet
        if c1 == self.list_category_code[0]:
            Type3_Report.copy_range(ws_main, ws, "C9:F34", dest_range)
            Type3_Report.copy_paste_cell(ws_main["C11"], cell_g10)
        elif c1 == self.list_category_code[1]:
            Type3_Report.copy_range(ws_main, ws, "G9:J34", dest_range)
            Type3_Report.copy_paste_cell(ws_main["G11"], cell_g10)
        elif c1 == self.list_category_code[2]:
            Type3_Report.copy_range(ws_main, ws, "K9:N34", dest_range)
            Type3_Report.copy_paste_cell(ws_main["K11"], cell_g10)
        elif c1 == self.list_category_code[3]:
            Type3_Report.copy_range(ws_main, ws, "O9:R34", dest_range)
            Type3_Report.copy_paste_cell(ws_main["O11"], cell_g10)
        elif c1 == self.list_category_code[4]:
            Type3_Report.copy_range(ws_main, ws, "S9:V34", dest_range)
            Type3_Report.copy_paste_cell(ws_main["S11"], cell_g10)
        elif c1 == self.list_category_code[5]:
            Type3_Report.copy_range(ws_main, ws, "W9:Z34", dest_range)
            Type3_Report.copy_paste_cell(ws_main["W11"], cell_g10)
        elif c1 == self.list_category_code[6]:
            Type3_Report.copy_range(ws_main, ws, "AA9:AD34", dest_range)
            Type3_Report.copy_paste_cell(ws_main["AA11"], cell_g10)

    def _put_data_NA(self, ws: Worksheet):
        """Fill in all the non available field rows with "Not Applicable" or "N/A" and color them gray.

        Args:
            ws (Worksheet): The worksheet to add data to.
        """
        start_row = 8
        end_row = 33

        for row in range(start_row, end_row + 1):
            if ws[f"G{row}"].value is None:
                # Set value and fill for each cell from A to J (1 to 10)
                for col in range(1, 11):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = self.gray_fill
                    if col == 2:
                        Type3_Report._set_cell_style(
                            cell=cell,
                            data="Not Applicable",
                            size=11,
                            bold=True,
                            horizontal=None,
                        )
                    elif col == 3 or col == 4:
                        Type3_Report._set_cell_style(
                            cell=cell, data="N/A", horizontal=None
                        )

    def _put_data_BCD(self, ws: Worksheet, c1: int):
        """Add data to columns: B, C, D from the instance dictionary variable.

        Args:
            ws (Worksheet): The worksheet to add data to.
            c1 (int): The category code.
        """
        start_row = 8
        for a1, a2 in self.dict_a1[c1].items():
            while ws[f"C{start_row}"].value is not None:
                start_row += 1
            # Write the value of a1 to column C starting from the 8th row
            Type3_Report._set_cell_style(
                cell=ws[f"C{start_row}"], data=a1, horizontal=None
            )
            # Write the value of a2[0] to column D starting from the 8th row
            Type3_Report._set_cell_style(
                cell=ws[f"D{start_row}"], data=a2[0].upper(), horizontal=None
            )
            # Write the value of a2[1] to column B starting from the 8th row
            Type3_Report._set_cell_style(
                cell=ws[f"B{start_row}"],
                data=a2[1],
                size=11,
                bold=True,
                horizontal=None,
            )

    @staticmethod
    def copy_range(
        source_sheet: Worksheet,
        target_sheet: Worksheet,
        source_range: str,
        target_range: str,
    ):
        """
        Copy the value of a range of cells from one sheet to another sheet's specified range.

        Args:
        source_sheet (Worksheet): The worksheet to copy data from.
        target_sheet (Worksheet): The worksheet to copy data to.
        source_range (str): The cell range in Excel format on the source sheet (e.g., 'A1:D4').
        target_range (str): The cell range in Excel format on the target sheet (e.g., 'A5:D8').
        """
        source_cells = list(source_sheet[source_range])
        target_cells = list(target_sheet[target_range])

        # Ensure that the source and target ranges are of the same size
        if len(source_cells) != len(target_cells) or any(
            len(src) != len(tgt) for src, tgt in zip(source_cells, target_cells)
        ):
            raise ValueError("Source and target ranges must be of the same size.")

        # Loop through the specified ranges
        for src_row, tgt_row in zip(source_cells, target_cells):
            for src_cell, tgt_cell in zip(src_row, tgt_row):
                # Copy value, font, alignment, fill
                tgt_cell.value = src_cell.value
                tgt_cell.font = Type3_Report.get_cell_font(src_cell)
                tgt_cell.alignment = Type3_Report.get_cell_alignment(src_cell)
                tgt_cell.fill = Type3_Report.get_cell_fill(src_cell)

    def _validate_category_sheet_input(
        self,
        wb: Workbook,
        category_sheet_name: str,
        method_name: str = "create_category_sheet",
    ):
        """Validate the input for creating a category sheet.

        Args:
            wb (Workbook): The workbook object to validate.
            category_sheet_name (str): The name of the category sheet to validate.
            method_name (str, optional): The method name where the validation happened. Defaults to "create_category_sheet".
        """
        Type3_Report.validate_wb(wb, method_name)
        Type3_Report.validate_str_list_tuple(category_sheet_name, method_name)

    @staticmethod
    def copy_paste_cell(
        src_cell: Cell,
        dest_cell: Cell,
        alt_value: bool = False,
        alt_font: bool = False,
        alt_alignment: bool = False,
        alt_fill: bool = False,
        alt_border: bool = False,
    ):
        """Copy and pasted the style and value of a cell to another cell.

        Args:
            src_cell (Cell): The cell to copy from.
            dest_cell (Cell): The cell to paste to.
            alt_value (bool, optional): False if not provide alternative value to set up. Defaults to False.
            alt_font (bool, optional): . Defaults to False.
            alt_alignment (bool, optional): _description_. Defaults to False.
            alt_fill (bool, optional): _description_. Defaults to False.
            alt_border (bool, optional): _description_. Defaults to False.
        """
        # Copy value and style of src_cell to dest_cell
        if not alt_value:
            dest_cell.value = src_cell.value
        if not alt_font:
            dest_cell.font = Type3_Report.get_cell_font(src_cell)
        if not alt_alignment:
            dest_cell.alignment = Type3_Report.get_cell_alignment(src_cell)
        if not alt_fill:
            dest_cell.fill = Type3_Report.get_cell_fill(src_cell)
        if not alt_border:
            dest_cell.border = Type3_Report.get_cell_border(src_cell)

    @staticmethod
    def get_cell_fill(cell: Cell) -> PatternFill:
        """Get the fill of a cell.

        Args:
            cell (Cell): The cell to get the fill from.

        Returns:
            PatternFill: The fill of the cell.
        """
        if cell.fill.fgColor.type == "rgb":
            fill = PatternFill(
                start_color=cell.fill.start_color,
                end_color=cell.fill.end_color,
                fill_type=cell.fill.fill_type,
            )
        else:
            console.log(f"Cell {cell.coordinate} does not have a fill color.")
        return fill

    @staticmethod
    def get_cell_border(cell: Cell) -> Border:
        """Get the border of a cell.

        Args:
            cell (Cell): The cell to get the border from.

        Returns:
            Border: The border of the cell.
        """
        return Border(
            left=Side(
                border_style=cell.border.left.border_style, color=cell.border.left.color
            ),
            right=Side(
                border_style=cell.border.right.border_style,
                color=cell.border.right.color,
            ),
            top=Side(
                border_style=cell.border.top.border_style, color=cell.border.top.color
            ),
            bottom=Side(
                border_style=cell.border.bottom.border_style,
                color=cell.border.bottom.color,
            ),
        )

    @staticmethod
    def get_cell_font(cell: Cell) -> Font:
        """Get the font of a cell.

        Args:
            cell (Cell): The cell to get the font from.

        Returns:
            Font: The font of the cell.
        """
        return Font(
            name=cell.font.name,
            size=cell.font.size,
            bold=cell.font.bold,
            italic=cell.font.italic,
            vertAlign=cell.font.vertAlign,
            underline=cell.font.underline,
            strike=cell.font.strike,
            color=cell.font.color,
        )

    @staticmethod
    def get_cell_alignment(cell: Cell) -> Alignment:
        """Get the alignment of a cell.

        Args:
            cell (Cell): The cell to get the alignment from.

        Returns:
            Alignment: The alignment of the cell.
        """
        return Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical,
            text_rotation=cell.alignment.text_rotation,
            wrap_text=cell.alignment.wrap_text,
            shrink_to_fit=cell.alignment.shrink_to_fit,
            indent=cell.alignment.indent,
        )

    def enable_links(self, wb: Workbook):
        """Enable links for all sheets in the workbook.

        Args:
            wb (Workbook): The workbook to enable links for.
        """
        try:
            # Enable links for all sheets in the workbook
            for ws in wb.worksheets:
                # Enable links for the main home sheet
                if ws.title == "Main Home":
                    self.link_main_home_sheet(ws)
                # Enable links for category home sheets
                elif "Home" in ws.title:
                    self.link_category_home_sheet(ws)
                # Enable links for drill down sheets
                elif "Drill" in ws.title:
                    self.link_drill_down_sheet(ws)
                # Enable links for field sheets
                elif "Fld" in ws.title:
                    self.link_field_sheet(ws)
        except Exception as e:
            console.log(f"Failed in enable_links(): {e}")
            raise e
        else:
            console.log("Links have been enabled for all sheets in the workbook.")

    def link_main_home_sheet(self, ws: Worksheet):
        """Enable links for the main home sheet.

        Args:
            ws (Worksheet): The main home sheet to enable links for.
        """
        # count is used to iterate through self.list_category_desc to get the target sheet name
        count = 0
        # Iterate through the merged cells that contain category names, make a step of 4 columns
        for col in range(
            Type3_Report.column_to_number("C"),
            Type3_Report.column_to_number("AD"),
            4,
        ):
            target_sheet = f"{self.list_category_desc[count][1]} Home"
            ws.cell(row=7, column=col).hyperlink = f"#'{target_sheet}'!A1"
            count += 1

    def link_category_home_sheet(self, ws: Worksheet):
        """Enable links for the category home sheets.

        Args:
            ws (Worksheet): The category home sheet to enable links for.
        """
        # Split the category home sheet name (e.g 'Cig Home' to 'Cig')
        cat_name = ws.title.split()[0]
        # Enable link for "Attribute Drill Down Listing" in cell G1
        Type3_Report.set_cell_link(
            source_sheet=ws, target_sheet=f"{cat_name} Drill", source_cell="G1"
        )
        # Enable link for "Main Summary" in cell J1
        Type3_Report.set_cell_link(
            source_sheet=ws, target_sheet="Main Home", source_cell="J1"
        )
        # Enable links for all the "Field #" in column A
        for row in range(self.skip_rows_cat_home_sheet, ws.max_row + 1):
            # Set the hyperlink if the cell has a default fill (gray fill indicates "Not Applicable")
            if Type3_Report.is_fill_default(ws[f"A{row}"]):
                Type3_Report.set_cell_link(
                    source_sheet=ws,
                    target_sheet=f"{cat_name} Fld {row - 7}",
                    source_cell=f"A{row}",
                )
            # Otherwise set the gray fill
            else:
                Type3_Report._set_cell_style(
                    cell=ws[f"A{row}"],
                    data=None,
                    size=11,
                    bold=True,
                    horizontal=None,
                    fill=self.gray_fill,
                )

    def link_drill_down_sheet(self, ws: Worksheet):
        """Enable links for the drill down sheets.

        Args:
            ws (Worksheet): The drill down sheet to enable links for.
        """
        # Split the drill down sheet name (e.g 'Cig Drill' to 'Cig')
        cat_name = ws.title.split()[0]
        # Enable link for "Category Home" in cell A4
        Type3_Report.set_cell_link(
            source_sheet=ws, target_sheet=f"{cat_name} Home", source_cell="A4"
        )

    def link_field_sheet(self, ws: Worksheet):
        """Enable links for the field sheets.

        Args:
            ws (Worksheet): The field sheet to enable links for.
        """
        # Split the field sheet name (e.g 'Cig Fld 1' to 'Cig')
        cat_name = ws.title.split()[0]
        # Enable link for "Category Home" in cell F1
        Type3_Report.set_cell_link(
            source_sheet=ws, target_sheet=f"{cat_name} Home", source_cell="F1"
        )

    @staticmethod
    def set_cell_link(
        source_sheet: Worksheet,
        target_sheet: str,
        source_cell: str,
        target_cell: str = "A1",
        style: bool = False,
    ):
        """Set a hyperlink from a source cell in a source sheet to a target cell in a target sheet.

        Args:
            source_sheet (Worksheet): The source sheet where the source cell is located.
            target_sheet (str): The target sheet where the hyperlink is pointed to.
            source_cell (str): The source cell where the hyperlink is set.
            target_cell (str, optional): The target cell in the target sheet. Defaults to "A1".
            style (bool, optional): False when not setting the 'Hyperlink' style of the cell. Defaults to False.
        """

        source_sheet[source_cell].hyperlink = f"#'{target_sheet}'!{target_cell}"
        if style:
            source_sheet[source_cell].style = "Hyperlink"

    @staticmethod
    def column_to_number(column_name: str) -> int:
        """Converts an Excel column name (e.g., 'A', 'B', ..., 'Z', 'AA', etc.) to its corresponding column number (e.g., 1, 2, ..., 26, 27, etc.).

        Args:
            column_name (str): The string column name to convert to a number.

        Returns:
            int: The corresponding column number.
        """
        # Validate input
        try:
            Type3_Report.validate_str_list_tuple(column_name, "column_to_number")

            column_number = 0
            for i, char in enumerate(reversed(column_name)):
                # Convert alphabet character to number, considering its position
                column_number += (ord(char) - ord("A") + 1) * (26**i)
            return column_number
        except Exception as e:
            console.log(f"Failed in column_to_number(): {e}")
            raise e

    def update_field_name(self, wb: Workbook):
        """Update the field sheet names based on the attribute names (column A) in the category home sheet.

        Args:
            wb (Workbook): The workbook object.
        """
        try:
            # Validate input
            Type3_Report.validate_wb(wb, "update_field_name")

            # Iterate over each category name in the list of category descriptions
            for category_name in self.list_category_desc:
                # Assemble the field sheet name and category sheet name based on the category name
                field_sheet = f"{category_name[1]} Fld"
                ws_cat = wb[f"{category_name[1]} Home"]

                # Iterate over each data row in category home sheet, that is, each attribute within the category starting from the last row
                for row in range(ws_cat.max_row, 7, -1):
                    # Iterate over each field sheet to find the one that corresponds to the attribute name
                    for ws in wb.worksheets:
                        # Find all the field sheets in that category (e.g. "Cig Fld" is in "Cig Fld 1", "Cig Fld 2", etc.)
                        if ws.title.startswith(f"{field_sheet} "):
                            # Find the row where the attribute name matches by checking column C in ws_cat
                            if ws_cat[f"C{row}"].value == ws["A6"].value.upper():
                                # Update the value in cell A6 (e.g. Category (1-4))
                                ws["A6"].value = (
                                    f"{ws_cat[f'B{row}'].value} ({ws_cat[f'E{row}'].value})"
                                )
                                # Change the field sheet name based on the value in column A in the field sheet
                                ws.title = (
                                    f"{field_sheet} {ws_cat[f'A{row}'].value.split()[-1]}"
                                )
        except Exception as e:
            console.log(f"Failed in update_field_name(): {e}")
            raise e
        else:
            console.log(
                "All field sheet names have been updated."
            )

    def reorder_sheets(self, wb: Workbook):
        """Reorder the sheets in the workbook based on the new order.

        Args:
            wb (Workbook): The workbook object.
        """
        try:
            new_order = self.get_sheet_order(wb)
            # Validate input
            self._validate_reorder_sheets(wb, new_order)
            # Create a new list of sheet objects based on the new_order
            wb._sheets = [wb[sheet_name] for sheet_name in new_order]
        except Exception as e:
            console.log(f"Failed in reorder_sheets(): {e}")
            raise e
        else:
            console.log("All sheets have been reordered successfully.")

    def get_sheet_order(self, wb: Workbook) -> list:
        """Get the new order of the sheets in the workbook based on the category descriptions.

        Args:
            wb (Workbook): The workbook object.

        Raises:
            ValueError: If the workbook is None.

        Returns:
            list: The new order of the sheets in the workbook.
        """
        # Validate input
        Type3_Report.validate_wb(wb, "get_sheet_order")
        # Initialize the new order list
        new_order = ["Main Home"]
        # Get all the sheetnames except "Main Home"
        sheet_names = [x for x in wb.sheetnames if x != "Main Home"]
        # Iterate through each category and attach category home sheet, dril down sheet, and field sheets in that order
        for cat in self.list_category_desc:
            new_order.append(f"{cat[1]} Home")
            new_order.append(f"{cat[1]} Drill")
            for sheet in sheet_names:
                if sheet.startswith(f"{cat[1]} Fld"):
                    new_order.append(sheet)
        # Return the new order list
        return new_order

    def _validate_reorder_sheets(
        self, wb: Workbook, new_order: list, method_name: str = "reorder_sheets"
    ):
        """Validate the input for the reorder_sheets() method.

        Args:
            wb (Workbook): The workbook object to validate.
            new_order (list): The new order of the sheets.
            method_name (str, optional): . Defaults to "reorder_sheets".

        Raises:
            ValueError: If the workbook is None or the sheet names in the new order do not exist in the workbook.
        """
        Type3_Report.validate_wb(wb, method_name)
        # Ensure all sheet names in the new_order list exist in the workbook
        if not all(sheet in wb.sheetnames for sheet in new_order):
            raise ValueError(
                f"Error in {method_name}(): One or more sheet names in the new order do not exist in the workbook."
            )

    def remove_template_sheets(self, wb: Workbook):
        """Remove the template sheets from the workbook.

        Args:
            wb (Workbook): The workbook object.

        Raises:
            ValueError: If the workbook is None.
        """
        try:
            # Validate input
            Type3_Report.validate_wb(wb, "remove_template_sheets")

            # Remove the template sheets
            for ws in wb.worksheets:
                if "_temp" in ws.title:
                    wb.remove(ws)
                    console.log(f"'{ws.title}' has been removed.")
        except Exception as e:
            console.log(f"Failed in remove_template_sheets(): {e}")
            raise e

    @staticmethod
    def validate_wb(wb: Workbook, method_name: str):
        """Validate the workbook object.

        Args:
            wb (Workbook): The workbook object to validate.
            method_name (str): The name of the method where the validation is performed.

        Raises:
            ValueError: If the workbook is None.
        """
        if wb is None:
            raise ValueError(f"Error in {method_name}(): The workbook is None.")

    @staticmethod
    def validate_ws(ws: Worksheet, method_name: str, check_title: str = None):
        """Validate the worksheet object.

        Args:
            ws (Worksheet): The worksheet object to validate.
            method_name (str): The name of the method where the validation is performed.

        Raises:
            ValueError: If the worksheet is None.
        """
        # Check if the worksheet is None
        if ws is None:
            raise ValueError(f"Error in {method_name}(): The worksheet is None.")
        # Check if the worksheet title is not the expected title
        if check_title is not None and ws.title != check_title:
            raise ValueError(
                f"Invalid input for {method_name}(): The worksheet title is not '{check_title}'."
            )

    @staticmethod
    def validate_df(df: pd.DataFrame, method_name: str):
        """Validate the DataFrame object.

        Args:
            df (pd.DataFrame): The DataFrame object to validate.
            method_name (str): The name of the method where the validation is performed.

        Raises:
            ValueError: If the DataFrame is None or empty.
        """
        if df is None or df.empty:
            raise ValueError(
                f"Invalid input for {method_name}(): The DataFrame is None or empty."
            )

    @staticmethod
    def validate_str_list_tuple(value, method_name: str):
        """Validate the input value to be a string, list, or tuple.

        Args:
            value (_type_): The value to validate, could be a string, list, or tuple.
            method_name (str): The name of the method where the validation is performed.

        Raises:
            ValueError: If the value is not a string, list, or tuple.
        """
        if Type3_Report.check_null_empty(value):
            raise ValueError(
                f"Invalid input for {method_name}(): value is None or empty."
            )
