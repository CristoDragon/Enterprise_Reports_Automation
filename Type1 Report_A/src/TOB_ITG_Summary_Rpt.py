import pandas as pd
import datetime
import numpy as np
import traceback
import os
import pdr.handlers.Console_Handler as console
import pdr.data.Dynamic_Report as dr
import pdr.period.ITG as itg
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell
from openpyxl import load_workbook
import zipfile



# Author: Dragon Xu (rxu@msa.com)
# Date: 07/22/2024
# Description: This class is used to run the Type 1 Report A job.


class TOB_ITG_Summary_Rpt:
    # Initialize all the instance variables
    def __init__(self, connection, table, report_id, template_path, output_path):
        # Initialize the connection to the Oracle database
        self.connection = connection
        # Initialize the table names
        self.tb_dr_reports = table
        # Initialize a list of report ID
        self.rid_cig_wdc = int(report_id[0])
        self.rid_ecig_wdc = int(report_id[1])
        self.rid_cgr_wdc = int(report_id[2])
        self.rid_cig = int(report_id[3])
        self.rid_ecig = int(report_id[4])
        self.rid_cgr = int(report_id[5])
        self.rid_otp = int(report_id[6])
        self.rid_otp_wdc = int(report_id[7])
        # Initialize the current datetime
        self.current_datetime = datetime.datetime.now().strftime('%m/%d/%Y %H:%M:%S')
        # Initialize the current week code
        self.cwk = itg.get_itg_period_code(self.connection)
        self.end_week = itg.get_itg_end_week(self.connection, self.cwk)
        # Initialize a list of week in format '%m/%d/%Y'
        self.weeks = self.init_weeks()
        # Initialize fills
        self.gray_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        # Initialize border styles
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")
        self.thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)
        # Initialize final deliverable sheet names
        self.out_final_sheets = ["Final_Comparison", "Comparison1", "Curr_DB", "Prev_DB"]
        self.out_summary_sheets = ["Summary", "cig", "blu", "cgr", "otp", "cigwdc", "bluwdc", "cgrwdc", "otpwdc"]
        # Initialize column names
        col1 = ["Concatenated", "Distributor Hierarchy", "Measures", "Manufacturer"]
        col2 = [f"CURRENT_{i}" for i in range(1, 157)]
        self.col_curr_db = col1 + col2
        self.col_fc_dates = ['RSD Vol','first date','last date','first fulldate','last fulldate','Date Range']
        self.col_summary = ["Distributor Name", "Customer Number", "Cigarettes", "e-Cigs", "Cigars", "Otp",
            "Additions", "Decreases", "Weeks Occurred", "Reason", 
            "Cigarettes (Ctns)", "e-Cigs (Units)", "Cigars (Sticks)", "Otp (Sticks)"
        ]
        self.calibri = "Calibri"
        # Initialize constants
        self.SPACE = ' '
        self.UNDERSCORE = '_'
        # Initialize template file names and output file names for dynamic reports
        self.dr_cig = "ITG_CIG_G360_XXXXX"
        self.dr_ecig = "ITG_ECIG_G360_XXXXX"
        self.dr_cgr = "ITG_CGR_G360_XXXXX"
        self.dr_otp = "ITG_OTP_G360_XXXXX"
        self.wdc_cig = "ITG_CIG_WDC_XXXXX"
        self.wdc_ecig = "ITG_ECIG_WDC_XXXXX"
        self.wdc_cgr = "ITG_CGR_WDC_XXXXX"
        self.wdc_otp = "ITG_OTP_WDC_XXXXX"
        if len(output_path) < 4:
            raise ValueError("Output path must contain 4 folders for CIG, ECIG, CGR, and OTP.")
        else:
            cig_folder = output_path[0]
            ecig_folder = output_path[1]
            cgr_folder = output_path[2]
            otp_folder = output_path[3]
        self.temp_dr_cig = os.path.join(template_path, f"{self.dr_cig}_Template.xlsx")
        self.out_dr_cig = os.path.join(cig_folder, f"{self.dr_cig}_{self.cwk}.xlsx")
        self.out_dr_ecig = os.path.join(ecig_folder, f"{self.dr_ecig}_{self.cwk}.xlsx")
        self.out_dr_cgr = os.path.join(cgr_folder, f"{self.dr_cgr}_{self.cwk}.xlsx")
        self.out_dr_otp = os.path.join(otp_folder, f"{self.dr_otp}_{self.cwk}.xlsx")
        self.temp_wdc = os.path.join(template_path, f"{self.wdc_cig}_Template.xlsx")
        self.out_wdc_cig = os.path.join(cig_folder, f"{self.wdc_cig}_{self.cwk}.xlsx")
        self.out_wdc_ecig = os.path.join(ecig_folder, f"{self.wdc_ecig}_{self.cwk}.xlsx")
        self.out_wdc_cgr = os.path.join(cgr_folder, f"{self.wdc_cgr}_{self.cwk}.xlsx")
        self.out_wdc_otp = os.path.join(otp_folder, f"{self.wdc_otp}_{self.cwk}.xlsx")
        # Initiaize template file names and output file names for final deliverables
        final_cig = "G360_XXXXX_Report"
        final_ecig = "G360_Ecig_XXXXX_Report"
        final_cgr = "G360_Cgr_XXXXX_Report"
        final_otp = "G360_OTP_XXXXX_Report"
        final_summary = "Volume_Change_Summary"
        self.temp_final_cig = os.path.join(template_path, "G360_XXXXX_Template.xlsx")
        self.temp_final_ecig = os.path.join(template_path, "G360_Ecigs_XXXXX_Template.xlsx")
        self.temp_final_cgr = os.path.join(template_path, "G360_Cgr_XXXXX_Template.xlsx")
        self.temp_final_otp = os.path.join(template_path, "G360_OTP_XXXXX_Template.xlsx")
        self.temp_final_summary = os.path.join(template_path, f"{final_summary}_Template.xlsx")
        self.out_final_cig = os.path.join(cig_folder, f"{final_cig}_{self.cwk}.xlsx")
        self.out_final_ecig = os.path.join(ecig_folder, f"{final_ecig}_{self.cwk}.xlsx")
        self.out_final_cgr = os.path.join(cgr_folder, f"{final_cgr}_{self.cwk}.xlsx")
        self.out_final_otp = os.path.join(otp_folder, f"{final_otp}_{self.cwk}.xlsx")
        self.out_final_summary = os.path.join(cig_folder, f"{final_summary}_{self.cwk}.xlsx")
        # Initialize input file namesll
        self.input_dr_cig = os.path.join(cig_folder, f"{self.dr_cig}_{self.cwk - 1}.xlsx")
        self.input_dr_ecig = os.path.join(ecig_folder, f"{self.dr_ecig}_{self.cwk - 1}.xlsx")
        self.input_dr_cgr = os.path.join(cgr_folder, f"{self.dr_cgr}_{self.cwk - 1}.xlsx")
        self.input_dr_otp = os.path.join(otp_folder, f"{self.dr_otp}_{self.cwk - 1}.xlsx")
        # Create a list of files to delete (we only want to keep the latest 3 weeks of data)
        # self.list_delete = [os.path.join(cig_folder, f"{self.dr_cig}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(ecig_folder, f"{self.dr_ecig}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(cgr_folder, f"{self.dr_cgr}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(otp_folder, f"{self.dr_otp}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(cig_folder, f"{self.wdc_cig}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(ecig_folder, f"{self.wdc_ecig}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(cgr_folder, f"{self.wdc_cgr}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(otp_folder, f"{self.wdc_otp}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(cig_folder, f"{final_cig}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(ecig_folder, f"{final_ecig}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(cgr_folder, f"{final_cgr}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(otp_folder, f"{final_otp}_{self.cwk - 3}.xlsx"),
        #                 os.path.join(cig_folder, f"{final_summary}_{self.cwk - 3}.xlsx")]
        
    def init_weeks(self) -> list:
        # Initialize an empty list to store the weeks
        weeks = []
        # Cutoff date for previous 155 weeks
        cutoff_date = self.end_week - datetime.timedelta(weeks=156)
        # Current date to calculate backwards from
        current_date = self.end_week
        # Calculate the weeks backwards from the current week
        while current_date >= cutoff_date:
            # Format and add current date to the list using f-string to avoid leading zero in month
            formatted_date = f"""{current_date.month}/{current_date.day}/{current_date.year}"""
            # Format and add current date to the list
            weeks.append(formatted_date)
            # Subtract one week for the next calculation
            current_date -= datetime.timedelta(weeks=1)
        return weeks
        
    @staticmethod
    def check_null_empty(value) -> bool:
        """
        Check if a value is None, empty, or contains no elements.

        Args:
            value (str, list, tuple): A value which can be a string, list, or tuple

        Returns:
            bool: True if the value is None, an empty string, or an empty collection, False otherwise
        """
        # Check if the value is None
        if value is None:
            return True
        # Check for string and use strip to remove any whitespace
        if isinstance(value, str):
            return value.strip() == ""
        # Check for list or tuple and if they are empty
        if isinstance(value, (list, tuple)):
            return len(value) == 0
        return False
    
    @staticmethod
    def validate_df(df: pd.DataFrame, df_name: str):
        """Validate the DataFrame object.

        Args:
            df (pd.DataFrame): The DataFrame object to validate.
            df_name (str): The name of the DataFrame object

        Raises:
            ValueError: If the DataFrame is None or empty.
        """
        if df is None or df.empty:
            console.log(f"Dataframe for '{df_name}' is None or empty.")
            # raise ValueError(f"DataFrame {df_name} is None or empty.")
        
    @staticmethod
    def dataframe_to_excel(
        df: pd.DataFrame,
        ws: Worksheet,
        skip_rows: int = 0,
        skip_cols: int = 0,
        sample_cell: Cell = None,
        alt_border: bool = False,
    ):
        """Copy the data from a DataFrame to a worksheet starting from the specified cell.

        Args:
            df (pd.DataFrame): The DataFrame to be copied to the worksheet
            ws (Worksheet): The worksheet to copy the DataFrame to
            skip_rows (int, optional): _description_. Defaults to 0.
            skip_cols (int, optional): _description_. Defaults to 0.
            sample_cell (Cell, optional): _description_. Defaults to None.
            alt_border (bool, optional): _description_. Defaults to False.
        """
        max_col = len(df.columns)
        max_row = len(df.index)
        for col in range(1, max_col + 1):
            for row in range(1, max_row + 1):
                col_complete = get_column_letter(col + skip_cols)
                row_complete = row + skip_rows
                dest_cell = ws[col_complete + str(row_complete)]
                # Copy the style of the sample cell to the destination cell if it is provided
                if sample_cell is not None:
                    TOB_ITG_Summary_Rpt.copy_paste_cell(
                        sample_cell, dest_cell, alt_value=True, alt_border=alt_border
                    )
                dest_cell.value = df.iat[row - 1, col - 1]
                
    @staticmethod
    def copy_paste_cell(
        src_cell: Cell,
        dest_cell: Cell,
        alt_value: bool = False,
        alt_font: bool = False,
        alt_alignment: bool = False,
        alt_fill: bool = False,
        alt_border: bool = False,
    ):
        """Copy and pasted the style and value of a cell to another cell.

        Args:
            src_cell (Cell): The cell to copy from.
            dest_cell (Cell): The cell to paste to.
            alt_value (bool, optional): False if not provide alternative value to set up. Defaults to False.
            alt_font (bool, optional): . Defaults to False.
            alt_alignment (bool, optional): _description_. Defaults to False.
            alt_fill (bool, optional): _description_. Defaults to False.
            alt_border (bool, optional): _description_. Defaults to False.
        """
        # Copy value and style of src_cell to dest_cell
        if not alt_value:
            dest_cell.value = src_cell.value
        if not alt_font:
            dest_cell.font = TOB_ITG_Summary_Rpt.get_cell_font(src_cell)
        if not alt_alignment:
            dest_cell.alignment = TOB_ITG_Summary_Rpt.get_cell_alignment(src_cell)
        if not alt_fill:
            dest_cell.fill = TOB_ITG_Summary_Rpt.get_cell_fill(src_cell)
        if not alt_border:
            dest_cell.border = TOB_ITG_Summary_Rpt.get_cell_border(src_cell)
             
    @staticmethod
    def get_cell_fill(cell: Cell) -> PatternFill:
        """Get the fill of a cell.

        Args:
            cell (Cell): The cell to get the fill from.

        Returns:
            PatternFill: The fill of the cell.
        """
        if cell.fill.fgColor.type == "rgb":
            fill = PatternFill(
                start_color=cell.fill.start_color,
                end_color=cell.fill.end_color,
                fill_type=cell.fill.fill_type,
            )
        else:
            console.log(f"Cell {cell.coordinate} does not have a fill color.")
        return fill

    @staticmethod
    def get_cell_border(cell: Cell) -> Border:
        """Get the border of a cell.

        Args:
            cell (Cell): The cell to get the border from.

        Returns:
            Border: The border of the cell.
        """
        return Border(
            left=Side(
                border_style=cell.border.left.border_style, color=cell.border.left.color
            ),
            right=Side(
                border_style=cell.border.right.border_style,
                color=cell.border.right.color,
            ),
            top=Side(
                border_style=cell.border.top.border_style, color=cell.border.top.color
            ),
            bottom=Side(
                border_style=cell.border.bottom.border_style,
                color=cell.border.bottom.color,
            ),
        )

    @staticmethod
    def get_cell_font(cell: Cell) -> Font:
        """Get the font of a cell.

        Args:
            cell (Cell): The cell to get the font from.

        Returns:
            Font: The font of the cell.
        """
        return Font(
            name=cell.font.name,
            size=cell.font.size,
            bold=cell.font.bold,
            italic=cell.font.italic,
            vertAlign=cell.font.vertAlign,
            underline=cell.font.underline,
            strike=cell.font.strike,
            color=cell.font.color,
        )

    @staticmethod
    def get_cell_alignment(cell: Cell) -> Alignment:
        """Get the alignment of a cell.

        Args:
            cell (Cell): The cell to get the alignment from.

        Returns:
            Alignment: The alignment of the cell.
        """
        return Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical,
            text_rotation=cell.alignment.text_rotation,
            wrap_text=cell.alignment.wrap_text,
            shrink_to_fit=cell.alignment.shrink_to_fit,
            indent=cell.alignment.indent,
        )
        
    @staticmethod
    def set_cell_style(
        cell,
        data,
        name="Arial",
        size=10,
        bold=False,
        underline=None,
        vertical="bottom",
        horizontal="center",
        wrapText=False,
        fill=None,
        border=None,
    ):
        try: 
            if data is not None:
                cell.value = data
            cell.font = Font(name=name, size=size, bold=bold, underline=underline)
            cell.alignment = Alignment(
                vertical=vertical, horizontal=horizontal, wrapText=wrapText
            )
            if fill is not None:
                cell.fill = fill
            if border is not None:
                cell.border = border
        except Exception as e:
            console.log(f"Error in set_cell_style(): {e}")
            raise e
        
    @staticmethod
    def close_wb(wb: Workbook, output_file_name: str):
        """Close and save the workbook with the formatted output file name

        Args:
            wb (Workbook): the workbook object
            output_file_name (str): the formatted output file name
        """
        try:
            # Save the workbook
            wb.save(f"{output_file_name}")
            wb.close()
        except Exception as e:
            console.log(f"Error when saving and closing the workbook: {e}")
            raise e
        
    @staticmethod
    def auto_adjust_column_width(ws: Worksheet, start_row: int = 1):
        """Auto adjust the column width of the worksheet based on the longest string of each column, starting from a specified row.

        Args:
            ws (Worksheet): The worksheet to adjust the column width.
            start_row (int): The first row to start scanning from (1-based index).
        """
        column_widths = {}
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                if cell.value:
                    # Update the width of the column if this cell's content is wider
                    column_width = len(str(cell.value))
                    if column_width > column_widths.get(cell.column_letter, 0):
                        column_widths[cell.column_letter] = column_width

        if ws.title == "Summary":
            column_widths["B"] = column_widths["B"] + 3

        # Adding a small buffer to column width for aesthetics
        buffer = 2
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width + buffer
    
    @staticmethod
    def delete_files(list_delete: list):
        """Delete the files in the list.

        Args:
            list_delete (list): The list of files to delete.
        """
        try:
            # Iterate through the list of files to delete
            for file_path in list_delete:
                # Check if file exist or not
                if os.path.exists(file_path):
                    # Remove the specified file path
                    os.remove(file_path)
                    console.log(f"File '{file_path}' has been deleted successfully.")
                else:
                    console.log(f"File '{file_path}' does not exist.")
        except FileNotFoundError:
            console.log(f"File {file_path} does not exist.")
        except PermissionError:
            console.log(f"Permission denied: unable to delete '{file_path}'.")
        except Exception as e:
            console.log(f"Error deleting the file {file_path}: {e}")
            

    
    def query_report_desc(self, report_id: int) -> str:
        """Query the dynamic report table in TOBM4P to get the report description for report_id.

        Args:
            report_id (int): The report ID to query the report description.

        Returns:
            str: The report description for the report_id.
        """
        try:
            query = f"SELECT report_desc FROM {self.tb_dr_reports} WHERE report_id = {report_id};"
            # Execute the query and get the data as a DataFrame
            df = pd.read_sql_query(query, self.connection)
            # Check if the DataFrame is None or empty
            TOB_ITG_Summary_Rpt.validate_df(df, self.tb_dr_reports)
            # Get the report description from the DataFrame
            return df["REPORT_DESC"][0]
        except Exception as e:
            console.log(f"Error in query_report_desc(): {e}")
            raise e
        
    def _put_data_dr_excel(self, ws: Worksheet, df: pd.DataFrame, report_id: int, data_name: str, row_count_cell: str):
        """Put the data from the DataFrame into the worksheet.

        Args:
            ws (Worksheet): The worksheet to put the data into.
            df (pd.DataFrame): The DataFrame to put into the worksheet.
            report_id (int): The report ID to get the report description.

        Raises:
            e: Any exception that occurs during the process.
        """
        try:
            # Update the report title on cell A1
            ws["A1"].value = f"WKLY REPORT - {data_name.upper()}"
            # Update the report subtitle on cell A2
            ws["A2"].value = f"{self.SPACE.join(data_name.replace(self.UNDERSCORE, self.SPACE).split()[:4])} CHANGES RPT"
            # Update the report creation date and time on cell A3
            ws["A3"].value = f"Report created on {self.current_datetime}"
            # Update the report details on cell A4
            ws["A4"].value = f"Generated from Report ID {report_id} on {self.tb_dr_reports.split(".")[1]}."
            # Update the row count at cell FD4
            ws[row_count_cell].value = int(df.shape[0])
            # Replace all the null values with 0
            df.fillna(0, inplace=True)
            # Put the dataframe into the worksheet starting from cell A8
            TOB_ITG_Summary_Rpt.dataframe_to_excel(df, ws, skip_rows=7)
        except Exception as e:
            console.log(f"Error in _put_data_dr_excel(): {e}")
            raise e
        
    def _set_style_dr_excel(self, ws: Worksheet, max_col: int, max_row: int):
        """Set the consistent cell style for the dynamic report.

        Args:
            ws (Worksheet): The worksheet to set the cell style.
            max_col (int): _description_
            max_row (int): _description_

        Raises:
            e: Any exception that occurs during the process.
        """
        try:
            # Set consistent cell style
            for row in ws.iter_rows(min_row=8, max_col=max_col, max_row=max_row):
                for cell in row:
                    TOB_ITG_Summary_Rpt.set_cell_style(cell, None, size=8, horizontal=None, wrapText=True, border=self.thin_border)
            # Auto adjust the column width of the worksheet
            TOB_ITG_Summary_Rpt.auto_adjust_column_width(ws, 7)
        except Exception as e:
            console.log(f"Error in _set_style_dr_excel(): {e}")
            raise e

    def create_dr_excel(self, df: pd.DataFrame, report_id: int, data_name: str, temp_file: str, out_file: str, row_count_cell: str):
        """Create a dynamic report in Excel format.

        Args:
            df (pd.DataFrame): the data to put into the dynamic report
            report_id (int): the report ID used to query the report data
            data_name (str): the name of the data
            temp_file (str): the template file path for the dynamic report
            out_file (str): the output file path for the dynamic report
            row_count_cell (str): _description_

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Create a workbook for the dynamic report template file
            wb = load_workbook(temp_file)
            # Get the active worksheet since there is only one worksheet in the template file
            ws = wb.active
            # Change the worksheet name
            ws.title = f"{data_name}_{self.cwk}"
            # Put the data into the worksheet
            self._put_data_dr_excel(ws, df, report_id, data_name, row_count_cell)
            # Set consistent cell style for the dynamic report
            self._set_style_dr_excel(ws, len(df.columns), ws.max_row)
            # Save and close the workbook
            TOB_ITG_Summary_Rpt.close_wb(wb, out_file)
        except Exception as e:
            console.log(f"Error in create_dr_excel(): {e}")
            raise e
        else:
            console.log(f"Dynamic report '{data_name}_{self.cwk}' has been created successfully.")
            
    def _init_current_sum(self, num_weeks: int) -> list:
        """Initialize the 'CURRENT-*' columns used in the query."""
        # Replace "CURRENT-0" with "CURRENT"
        current_sum = [f"""SUM(CASE WHEN FPDS.PERIOD_CODE = {self.cwk - i} THEN TOT_VOL END) AS \"{('CURRENT' if i == 0 else f'CURRENT-{i}')}\"""" for i in range(0, 156)]
        # Return ths list of 'CURRENT-*' columns
        return current_sum
    
    def get_dr_data(self, report_id: int) -> pd.DataFrame:
        # Query the database to get dynamic report data
        df = dr.collect_dynamic_report(self.connection, self.tb_dr_reports, report_id)
        # Check if the dataframe is None or empty
        TOB_ITG_Summary_Rpt.validate_df(df, self.dr_cig)
        return df
        
    def get_dr_wdc(self, report_id: int, df_name: str) -> pd.DataFrame:
        # Query the database to get dynamic report data
        df = dr.collect_dynamic_report(self.connection, self.tb_dr_reports, report_id)
        # Check if the dataframe is None or empty
        TOB_ITG_Summary_Rpt.validate_df(df, df_name)
        return df
    
    def _update_weeks_curr_db(self, ws: Worksheet, weeks: list):
        """Update the week codes in the worksheet for the Curr_DB.

        Args:
            ws (Worksheet): The worksheet to update the week codes.
            weeks (list): The list of week codes to update in the worksheet.
        """
        # Update all the weeks in row 2 starting from column E
        for i, week in enumerate(weeks):
            week_cell = ws.cell(row=2, column=i+5)
            # Set the value and style of the cell
            TOB_ITG_Summary_Rpt.set_cell_style(week_cell, data=week, vertical="center", bold=True, wrapText=True)
                
    def _process_df_curr_db(self, df: pd.DataFrame, col_filter: str) -> pd.DataFrame:
        """Process the dataframe for the Curr_DB worksheet.

        Args:
            df (pd.DataFrame): The dataframe to process.
            col_filter (str): The filter string used to filter the dataframe.

        Returns:
            pd.DataFrame: Any exception raised during the process.
        """
        # Process the dataframe by filterting for the rows with "Ctns" in column 'Measures'
        df = df[df["Measures"] == col_filter]
        # Drop the column 'Sort Order'
        if "Sort Order" in df.columns:
            df.drop(columns=["Sort Order"], inplace=True)
        # Create a new column by concatenating the first three columns
        df[self.col_curr_db[0]] = df.iloc[:, 0].astype(str) + df.iloc[:, 1].astype(str) + df.iloc[:, 2].astype(str)
        # Fill all the null values with 0
        df.fillna(0, inplace=True)
        # Rearrange columns to make the new column the first one
        cols = df.columns.tolist() 
        # Move the last column (new column) to the first position 
        cols = [cols[-1]] + cols[:-1]
        # Reassign the columns to the dataframe
        df = df[cols]
        return df
    
    def _set_style_curr_db(self, ws: Worksheet, max_row: int):
        """Set the consistent cell style for the Curr_DB worksheet.

        Args:
            ws (Worksheet): The worksheet to set the cell style.
            max_row (int): The maximum row number in the worksheet.
        """
        # Set consistent cell style for columns B, C, D
        for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=2, max_col=4):
            for cell in row:
                TOB_ITG_Summary_Rpt.set_cell_style(cell, None, vertical="center", horizontal="left", bold=True, fill=self.gray_fill, border=self.thin_border)
        # Auto adjust the column width of the worksheet
        #TOB_ITG_Summary_Rpt.auto_adjust_column_width(ws, 2)
        
    
    def create_ws_curr_db(self, wb: Workbook, ws_name: str, df: pd.DataFrame, weeks: list, col_filter: str):
        """Create the Curr_DB worksheet in the workbook.

        Args:
            wb (Workbook): The workbook to create the worksheet in.
            ws_name (str): The name of the template worksheet.
            df (pd.DataFrame): The dataframe to put into the worksheet.
            weeks (list): The list of week codes to update in the worksheet.
            col_filter (str): The filter string used to filter the dataframe.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Load the worksheet Curr_DB from the template file
            ws = wb[ws_name]
            # Update all the weeks in row 2 starting from column E
            self._update_weeks_curr_db(ws, weeks)
            # Process the dataframe by filtering and creating a new column
            df_processed = self._process_df_curr_db(df, col_filter)
            # Put the dataframe into the worksheet starting from cell B3
            TOB_ITG_Summary_Rpt.dataframe_to_excel(df_processed, ws, skip_rows=2)
            # Set consistent cell style for the worksheet
            self._set_style_curr_db(ws, max_row=df_processed.shape[0] + 2)
        except Exception as e:
            console.log(f"Error in create_ws_curr_db(): {e}")
            raise e
        else:
            console.log(f"Worksheet '{ws_name}' has been created successfully.")
            
    def create_ws_prev_db(self, wb: Workbook, ws_name: str, weeks: list, col_filter: str, input_dr: str):
        """Create the Prev_DB worksheet in the workbook.

        Args:
            wb (Workbook): The workbook to create the worksheet in.
            ws_name (str): The name of the template worksheet.
            weeks (list): The list of week codes to update in the worksheet.
            col_filter (str): The filter string used to filter the dataframe.
            input_dr (str): The input file path for the previous week's dynamic report.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Get data from the previous week's dynamic report, skip the first 6 rows of the dataframe
            df_input = pd.read_excel(input_dr, header=0, skiprows=6)
            #df_input = pd.read_csv(input_dr, delimiter='|')
            # Create the worksheet Prev_DB using previous week's data
            self.create_ws_curr_db(wb, ws_name, df_input, weeks, col_filter)
        except Exception as e:
            console.log(f"Error in create_ws_prev_db(): {e}")
            raise e
        
    def _merge_df_curr_prev(self, wb: Workbook) -> pd.DataFrame:
        """Merge the dataframes from the Curr_DB and Prev_DB worksheets.

        Args:
            wb (Workbook): The workbook to get the data from.

        Raises:
            e: Any exception raised during the process.

        Returns:
            pd.DataFrame: the merged dataframe.
        """
        try:
            # Get the data from Curr_DB and Prev_DB worksheets as datafrmes
            data1 = wb[self.out_final_sheets[2]].values
            data2 = wb[self.out_final_sheets[3]].values
            # Skip the first two rows of worksheet
            next(data1)
            next(data1)
            next(data2)
            next(data2)
            # Create the dataframes from the worksheet values
            df_prev = pd.DataFrame(data2, columns=self.col_curr_db)
            df_curr = pd.DataFrame(data1, columns=self.col_curr_db)
            # Join the two dataframes on the first column 'Concatenated'
            df_merged = pd.merge(df_curr, df_prev, on=self.col_curr_db[0], how="inner", suffixes=('_Curr', '_Prev'))
            # Replace all the null values with 0
            df_merged.fillna(0, inplace=True)
            return df_merged
        except Exception as e:
            console.log(f"Error in _merge_df_curr_prev(): {e}")
            raise e
    
    def _process_df_merged(self, df_merged: pd.DataFrame, cols: list, end_week: int) -> pd.DataFrame:
        """Process the merged dataframe to calculate the differences and sums.

        Args:
            df_merged (pd.DataFrame): The merged dataframe to process.
            cols (list): The list of column names to create in the dataframe.
            end_week (int): _description_

        Raises:
            e: Any exception raised during the process.

        Returns:
            pd.DataFrame: the processed dataframe.
        """
        try:
            # Calculate the difference between the CURRENT_* columns
            for i in range(1, end_week):
                df_merged[f"C Diff {i}"] = df_merged[f"CURRENT_{i + 1}_Curr"].astype(int) - df_merged[f"CURRENT_{i}_Prev"].astype(int)
            # Calculate the sum of the next 13 CURRENT columns
            df_merged[cols[0]] = df_merged.iloc[:, 17:17+end_week].sum(axis=1)
            df_merged[cols[1]] = df_merged.iloc[:, 175:175+end_week].sum(axis=1)
            # Calculate the difference between the sum of the CURRENT columns
            df_merged[cols[2]] = df_merged[cols[0]] - df_merged[cols[1]]
            # Calculate the ABS sum of change
            df_merged[cols[6]] = abs(df_merged[cols[2]]) + df_merged.filter(like="C Diff").applymap(abs).sum(axis=1)
            return df_merged
        except Exception as e:
            console.log(f"Error in _process_df_merged(): {e}")
            raise e  
        
    def _process_df_filtered(self, df_filtered: pd.DataFrame, cols: list, filter_str: list, end_week: int) -> pd.DataFrame:
        """Process the filtered dataframe to create new columns and filter the data.
        
        Args:
            df_filtered (pd.DataFrame): The filtered dataframe to process.
            cols (list): The list of column names to create in the dataframe.
            filter_str (list): The filter string used to filter the dataframe.
            end_week (int): _description_
            
            Raises:
            e: Any exception raised during the process.
            
            Returns:
            pd.DataFrame: the processed dataframe.
        """
        try:
            # Update all the C Diff columns to include *
            for i in range(1, end_week):
                df_filtered[f"C Diff {i}"] = df_filtered.apply(
                    lambda row: (
                        "*   " if row[f"CURRENT_{i + 1}_Curr"] == 0 else "") +
                        f"""{int(row[f"CURRENT_{i + 1}_Curr"]) - int(row[f"CURRENT_{i}_Prev"])}""" +
                        ("   *" if row[f"CURRENT_{i}_Prev"] == 0 else ""),
                    axis=1
                )
            # Calculate the AVG volume for the first 13 CURRENT columns 
            df_filtered[cols[5]] = df_filtered.iloc[:, 5:18].mean(axis=1)
            # Create the column 'Dist/Packing' by concatenating the 2, 3, 4 column with '/'
            df_filtered[cols[4]] = df_filtered.apply(
                    lambda row: (row['Distributor Hierarchy_Curr']) + " / " + (row['Measures_Curr']) + " / " + (row["Manufacturer_Curr"]), axis=1)
            # Create the column 'Dist Name / Cust #'
            df_filtered[cols[3]] = df_filtered.apply(
                    lambda row: (row['Distributor Hierarchy_Curr'].split(self.SPACE, maxsplit=1)[1].title()) + " (" + (row['Distributor Hierarchy_Curr'].split(self.SPACE)[0]) + ")", axis=1)
            # Filter the dataframe based on the filter string
            if filter_str is not None:
                if len(filter_str) == 1:
                    df_filtered = df_filtered[df_filtered[cols[4]].str.contains(filter_str[0], regex=True)]
                if len(filter_str) == 2:
                    df_filtered = df_filtered[df_filtered[cols[4]].str.contains(filter_str[0], regex=True) | df_filtered[cols[4]].str.contains(filter_str[1], regex=True)]
            # Select the columns to keep
            col_final = cols[3:] + [f"C Diff {i}" for i in range(1, end_week)] + [cols[2]]
            # Return the final dataframe
            return df_filtered[col_final]
        except Exception as e:
            console.log(f"Error in _process_df_filtered(): {e}")
            raise e
        
        
    def _set_style_comparison(self, ws: Worksheet, max_row: int, end_week: int):
        """Set the consistent cell style for the Comparison worksheet.

        Args:
            ws (Worksheet): The worksheet to set the cell style.
            max_row (int): The maximum row number in the worksheet.
            end_week (int): _description_
        """
        # Set consistent cell style for the first 4 columns
        for row in ws.iter_rows(min_row=6, max_row=max_row, min_col=1, max_col=4):
            for cell in row:
                TOB_ITG_Summary_Rpt.set_cell_style(cell, None, name=self.calibri, size=11, horizontal=None)
        # Set consistent cell style for the rest of colunns
        for row in ws.iter_rows(min_row=6, max_row=max_row, min_col=5, max_col=4+end_week):
            for cell in row:
                TOB_ITG_Summary_Rpt.set_cell_style(cell, None,  name=self.calibri, size=11)
        # Set the number format of the cell to show as an integer
        for i in range(6, ws.max_row + 1):
            ws.cell(row=i, column=3).number_format = '0'
        # Set filter for the header row
        ws.auto_filter.ref = f"A5:R{ws.max_row}"
        # Auto adjust the column width of the worksheet
        #TOB_ITG_Summary_Rpt.auto_adjust_column_width(ws, 5)
        
    def _update_weeks_comparison(self, ws: Worksheet, end_week: int):
        """Update the week codes in the worksheet Final_Comparison.

        Args:
            ws (Worksheet): The worksheet to update the week codes.
            end_week (int): _description_
        """
        # Update the week code at row 5 starting from column E
        for i, week in enumerate(self.weeks[1:end_week]):
            week_cell = ws.cell(row=5, column=i+5)
            value = f"{self.cwk - 1 - i}     {week}"
            # Set the value and style of the cell
            TOB_ITG_Summary_Rpt.set_cell_style(week_cell, data=value, name=self.calibri, size=11, wrapText=True)
        # Update the the last cell at row 5
        last_cell = ws.cell(row=5, column=end_week + 4)
        value1 = f"{self.cwk-end_week} - {self.cwk-end_week-12}  ABS Chngs"
        TOB_ITG_Summary_Rpt.set_cell_style(last_cell, data=value1, name=self.calibri, size=11, wrapText=True)
        # Update the report creation time on cell A2
        ws["A2"].value = f"Report created on {self.current_datetime}"
        
    def create_ws_comparison(self, wb: Workbook, ws_name: str, filter_str: list, end_week: int):
        """Create the Final_Comparison worksheet in the workbook.

        Args:
            wb (Workbook): The workbook to create the worksheet in.
            ws_name (str): The name of the template worksheet.
            filter_str (list): The filter string used to filter the dataframe.
            end_week (int): _description_

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Load the worksheet Final_Comparison from the template file
            ws = wb[ws_name]
            # Update the week code at row 5 starting from column E
            self._update_weeks_comparison(ws, end_week)
            # Merge the data from Curr_DB and Prev_DB worksheets
            df_merged = self._merge_df_curr_prev(wb)
            # Initialize column names
            cols = ["Sum_Change_Curr", "Sum_Change_Prev", "Sum_Change_Diff", "Dist Name / Cust #", "Dist/Packing", "AVG_Volume", "ABS_Sum_Change"]
            # Process the merged dataframe by creating new columns for ABS_Sum_change and Sum_Change_Diff
            df_merged = self._process_df_merged(df_merged, cols, end_week)       
            # Filter for the rows with non-zero values in the CURRENT columns
            df_filtered = df_merged[df_merged[cols[6]] != 0]
            # Process the filtered dataframe and return the final dataframe
            df_final = self._process_df_filtered(df_filtered, cols, filter_str, end_week)
            # Sort the dataframe by Totale ABS change
            df_final.sort_values(by=cols[6], ascending=False, inplace=True)
            # Put the dataframe into the worksheet starting from cell A6
            TOB_ITG_Summary_Rpt.dataframe_to_excel(df_final, ws, skip_rows=5)
            # Set consistent cell style for the worksheet
            self._set_style_comparison(ws, df_final.shape[0] + 5, end_week)
        except Exception as e:
            console.log(f"Error in create_ws_comparison(): {e}")
            raise e
        else:
            console.log(f"Worksheet '{ws_name}' has been created successfully.")
        

    def create_final_excel(self, df: pd.DataFrame, col_filter: str, temp_final: str, input_dr: str, out_final: str, filter_str: list, end_week: int = 14):
        """Create the final deliverable in Excel format.

        Args:
            df (pd.DataFrame): The dataframe to put into the worksheet Curr_DB.
            col_filter (str): The filter string used to filter the dataframe.
            temp_final (str): The template file path for the final deliverable.
            input_dr (str): The input file path for the previous week's dynamic report.
            out_final (str): The output file path for the final deliverable.
            filter_str (list): The filter string used to filter the dataframe.
            end_week (int, optional): The number of week columns to keep. Defaults to 14.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Load the workbook for the final deliverable template file
            wb = load_workbook(temp_final)
            # Create the worksheet Curr_DB
            self.create_ws_curr_db(wb, self.out_final_sheets[2], df, self.weeks[:len(self.weeks) - 1], col_filter)
            # Create the worksheet Prev_DB
            self.create_ws_prev_db(wb, self.out_final_sheets[3], self.weeks[1:], col_filter, input_dr)
            # Create the worksheet Final_Comparison
            self.create_ws_comparison(wb, self.out_final_sheets[0], filter_str, end_week) 
            # Save and close the workbook
            TOB_ITG_Summary_Rpt.close_wb(wb, out_final)
        except Exception as e:
            console.log(f"Error in create_final_excel(): {e}")
            raise e
        else:
            console.log(f"Final deliverable '{out_final}' has been created successfully.")
            
           
                     
    def cig_job(self):
        try:
            # Query the database to get dynamic report data
            df_g360 = self.get_dr_data(self.rid_cig)
            # Create an excel file for the dynamic report data
            self.create_dr_excel(df_g360, self.rid_cig, self.dr_cig, self.temp_dr_cig, self.out_dr_cig, "FD4")
            # Create an excel file for the final deliverable
            self.create_final_excel(df_g360, "Ctns", self.temp_final_cig, self.input_dr_cig, self.out_final_cig, ["ITG Brands"])
            # Query the database to get WDC data
            df_wdc = self.get_dr_wdc(self.rid_cig_wdc, self.wdc_cig)
            # Create an excel file for the WDC data
            self.create_dr_excel(df_wdc, self.rid_cig_wdc, self.wdc_cig, self.temp_wdc, self.out_wdc_cig, "G4")
        except Exception as e:
            console.log(f"Error in cig_job(): {e}")
            raise e
        else:
            console.log("CIG part has completed successfully.")
            
        
    def ecig_job(self):
        try:
            # Query the database to get dynamic report data
            df_g360 = self.get_dr_data(self.rid_ecig)
            # Create a excel file for the dynamic report data
            self.create_dr_excel(df_g360, self.rid_ecig, self.dr_ecig, self.temp_dr_cig, self.out_dr_ecig, "FD4")
            # Create an excel file for the final deliverable
            self.create_final_excel(df_g360, "Units", self.temp_final_ecig, self.input_dr_ecig, self.out_final_ecig, ["blu ecigs"])
            # Query the database to get WDC data
            df_wdc = self.get_dr_wdc(self.rid_ecig_wdc, self.wdc_ecig)
            # Create an excel file for the WDC data
            self.create_dr_excel(df_wdc, self.rid_ecig_wdc, self.wdc_ecig, self.temp_wdc, self.out_wdc_ecig, "G4")
        except Exception as e:
            console.log(f"Error in ecig_job(): {e}")
            raise e
        else:
            console.log("ECIG part has completed successfully.")
            
            
    def cgr_job(self):
        try:
            # Query the database to get dynamic report data
            df_g360 = self.get_dr_data(self.rid_cgr)
            # Create a excel file for the dynamic report data
            self.create_dr_excel(df_g360, self.rid_cgr, self.dr_cgr, self.temp_dr_cig, self.out_dr_cgr, "FD4")
            # Create an excel file for the final deliverable. Note that we use \ to escape the parentheses in the filter string.
            self.create_final_excel(df_g360, "Vol", self.temp_final_cgr, self.input_dr_cgr, self.out_final_cgr, ["ITG Cigars Inc \\(Mmc\\)"])
            # Query the database to get WDC data
            df_wdc = self.get_dr_wdc(self.rid_cgr_wdc, self.wdc_cgr)
            # Create an excel file for the WDC data
            self.create_dr_excel(df_wdc, self.rid_cgr_wdc, self.wdc_cgr, self.temp_wdc, self.out_wdc_cgr, "G4")
        except Exception as e:
            console.log(f"Error in cgr_job(): {e}")
            raise e
        else:
            console.log("CGR part has completed successfully.")
            
    def otp_job(self):
        try:
            # Query the database to get dynamic report data
            df_g360 = self.get_dr_data(self.rid_otp)
            # Create a excel file for the dynamic report data
            self.create_dr_excel(df_g360, self.rid_otp, self.dr_otp, self.temp_dr_cig, self.out_dr_otp, "FD4")
            # Create an excel file for the final deliverable.
            self.create_final_excel(df_g360, "Vol", self.temp_final_otp, self.input_dr_otp, self.out_final_otp, ["Modern Oral", "Wraps"])
            # Query the database to get WDC data
            df_wdc = self.get_dr_wdc(self.rid_otp_wdc, self.wdc_otp)
            # Create an excel file for the WDC data
            self.create_dr_excel(df_wdc, self.rid_otp_wdc, self.wdc_otp, self.temp_wdc, self.out_wdc_otp, "G4")
        except Exception as e:
            console.log(f"Error in otp_job(): {e}")
            raise e
        else:
            console.log("CGR part has completed successfully.")
            
    def _put_wdc_data(self, ws: Worksheet, input_wdc: str):
        """Put the WDC data into the WDC worksheet in final deliverable summary report.

        Args:
            ws (Worksheet): The worksheet to put the data into.
            input_wdc (str): The input file path to get the WDC data.
        """
        # Read wdc report data from excel for each category
        df = pd.read_excel(input_wdc, header=0, skiprows=6)
        # Check if the DataFrame is None or empty
        TOB_ITG_Summary_Rpt.validate_df(df, input_wdc)
        # Concatenate the 2nd and 5th columns to create a new column
        df[self.col_curr_db[0]] = df.iloc[:, 1].astype(str) + df.iloc[:, 4].astype(str)
        # Reorder the columns so that the Concatenated column is the first column
        cols = df.columns.tolist()
        cols = [cols[-1]] + cols[:-1]
        df = df[cols]
        # Put the data into the worksheet starting from the second row and first column
        TOB_ITG_Summary_Rpt.dataframe_to_excel(df, ws, skip_rows=1)
        
    def _set_wdc_style(self, ws: Worksheet):
        """Set the consistent cell style for the WDC worksheet in final deliverable summary report.

        Args:
            ws (Worksheet): The worksheet to set the cell style.
        """
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
            for cell in row:
                # Set consistent style for column 1, 3, 4, 8
                if cell.column in [1, 3, 4, 8]:
                    TOB_ITG_Summary_Rpt.set_cell_style(cell, None, name=self.calibri, size=11, horizontal=None)
                # Set consistent style for the rest of the columns
                else:
                    TOB_ITG_Summary_Rpt.set_cell_style(cell, None, name=self.calibri, size=11, horizontal="right")
        # Automatically adjust the column width of the worksheet
        TOB_ITG_Summary_Rpt.auto_adjust_column_width(ws)
                    
         
    def _create_wdc_sheet(self, wb: Workbook, input_wdc: str, sheet_wdc: str):
        """Create the WDC worksheet in the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to create the worksheet in.
            input_wdc (str): The input file path to get the WDC data.
            sheet_wdc (str): The name of the worksheet to create.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Load the worksheet for wdc report
            ws = wb[sheet_wdc]
            # Put data into the worksheet
            self._put_wdc_data(ws, input_wdc)
            # Set the style of the worksheet
            self._set_wdc_style(ws)
        except Exception as e:
            console.log(f"Error in _create_wdc_sheet(): {e}")
            raise e
        else:
            console.log(f"Worksheet '{sheet_wdc}' has been created successfully.")
        
            
    def _create_wdc_sheets(self, wb: Workbook):
        """Create the WDC worksheets for all categories in the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to create the worksheets in.

        Raises:
            ValueError: _description_
            e: Any exception raised during the process.
        """
        try:
            # Initialize two lists for wdc file path and output wdc sheet names
            paths = [self.out_wdc_cig, self.out_wdc_ecig, self.out_wdc_cgr, self.out_wdc_otp]
            sheets = self.out_summary_sheets[5:]
            # Check if these two lists have the same length
            if len(paths) != len(sheets):
                raise ValueError("The number of wdc file paths and output wdc sheet names do not match.")
            # Loop through all the categories to create wdc worksheets
            for input_wdc, sheet_wdc in zip(paths, sheets):
                self._create_wdc_sheet(wb, input_wdc, sheet_wdc)
        except Exception as e:
            console.log(f"Error in _create_wdc_sheets(): {e}")
            raise e
        else:
            console.log("All of the WDC worksheets have been created successfully.")
            
    def add_new_columns(self, row):
        # Filter the row for the date columns and drop zero values
        filtered_row = row[self.col_dates].replace(0, pd.NA).dropna()
        if not filtered_row.empty:
            # First non-zero date
            first_date = filtered_row.index[0]
            # Last non-zero date
            last_date = filtered_row.index[-1]
        else:
            # Handle case where there are no non-zero values
            first_date, last_date = pd.NA, pd.NA
        return pd.Series([first_date, last_date], index=['first full', 'last full'])
        
            
    def _put_fc_data(self, ws: Worksheet, input_fc: str):
        try:
            # Read final comparison data from excel for each category
            df = pd.read_excel(input_fc, header=0, skiprows=4)
            # Put the old data into the worksheet starting from A6
            TOB_ITG_Summary_Rpt.dataframe_to_excel(df, ws, skip_rows=5)
            # Remove any white space and special character * from column 5 to 18
            for column in df.columns[4:18]:
                df[column] = df[column].astype(str).str.replace(r'[\s*]', '', regex=True)
            # Convert column 4 ~ 18 to int
            df.iloc[:, 4:18] = df.iloc[:, 4:18].astype(int)
            # Initialize column names
            rsd_vol = self.col_fc_dates[0]
            first_date = self.col_fc_dates[1]
            last_date = self.col_fc_dates[2]
            first_fulldate = self.col_fc_dates[3]
            last_fulldate = self.col_fc_dates[4]
            # Create a new column 'RSD Vol' that sums from the 5th column to 18th column
            df[rsd_vol] = df.iloc[:, 4:18].sum(axis=1)
            # Convert that column to int value
            df[rsd_vol] = df[rsd_vol].astype(int)
            # Get the column names from 5th column to 17 column and store it as a list
            self.col_dates = df.columns[4:17].tolist()
            # Create the rest of new columns related to first date and last date
            new_row = df.apply(self.add_new_columns, axis=1)
            # Join
            df = df.join(new_row)
            # Replace all the null values with empty string so that the split is successful
            df['first full'] = df['first full'].fillna(' ')
            df['last full'] = df['last full'].fillna(' ') 
            # Split the column
            df[[first_date, first_fulldate]] = df['first full'].str.split(expand=True)
            df[[last_date, last_fulldate]] = df['last full'].str.split(expand=True)
            # Safely convert to int, replacing non-convertible values with NaN
            df[first_date] = pd.to_numeric(df[first_date], errors='coerce')
            df[last_date] = pd.to_numeric(df[last_date], errors='coerce')
            # Strip white spaces, considering null values
            df[first_fulldate] = df[first_fulldate].str.strip().replace('', None)
            df[last_fulldate] = df[last_fulldate].str.strip().replace('', None)
            # Create a new column 'Date Range' with null handling
            df[self.col_fc_dates[5]] = df.apply(
                lambda row: None if pd.isnull(row[first_fulldate]) or pd.isnull(row[last_fulldate]) 
                else (row[first_fulldate] if row[first_fulldate] == row[last_fulldate] 
                    else f"{row[last_fulldate]} - {row[first_fulldate]}"), 
                axis=1
            )
            # Reorder the columns and only keep the dates columns that are newly created
            cols_reorder = self.col_fc_dates
            df_dates = df[cols_reorder]
            # Put the new data into the worksheet starting from S6
            TOB_ITG_Summary_Rpt.dataframe_to_excel(df_dates, ws, skip_rows=5, skip_cols=18)
            # Update the period code in the headers of the worksheet, end_week parameter is default to 14 for all categories
            self._update_weeks_comparison(ws, 14)
        except zipfile.BadZipFile:
            console.log(f"Error: The file '{input_fc}' is corrupted or being occupied by other process. Verify the file and try run the program again.")
        except FileNotFoundError:
            console.log(f"The file '{input_fc}' was not found. Ensure the path is correct.")
        except Exception as e:
            console.log(f"Error in _put_fc_data(): {e}")
            raise e
        
        
    
    def _set_fc_style(self, ws: Worksheet):
        """Set the consistent cell style for the FC worksheet in final deliverable summary report.

        Args:
            ws (Worksheet): The worksheet to set the cell style.
        """
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=18):
            for cell in row:
                # Set consistent style for column 1 ~ 4
                if cell.column in [1, 2, 3, 4]:
                    TOB_ITG_Summary_Rpt.set_cell_style(cell, None, name=self.calibri, size=11, horizontal=None)
                # Set consistent style for the rest of the columns
                else:
                    TOB_ITG_Summary_Rpt.set_cell_style(cell, None, name=self.calibri, size=11)
        # Set the number format of the cell to show as an integer
        for i in range(6, ws.max_row + 1):
            ws.cell(row=i, column=3).number_format = '0'
        # Automatically adjust the column width of the worksheet
        #TOB_ITG_Summary_Rpt.auto_adjust_column_width(ws, 5)
            
    def _create_fc_sheet(self, wb: Workbook, input_fc: str, sheet_fc: str):
        """Create the FC worksheet in the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to create the worksheet in.
            input_fc (str): The input file path to get the FC data.
            sheet_fc (str): The name of the worksheet to create.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Load the worksheet for wdc report for each category
            ws = wb[sheet_fc]
            # Put data into the worksheet
            self._put_fc_data(ws, input_fc)
            # Set the style of the worksheet
            self._set_fc_style(ws)
        except Exception as e:
            console.log(f"Error in _create_fc_sheet(): {e}")
            raise e
        else:
            console.log(f"Worksheet '{sheet_fc}' has been created successfully.")
            
    def _create_fc_sheets(self, wb: Workbook):
        """Create the FC worksheets for all categories in the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to create the worksheets in.

        Raises:
            ValueError: _description_
            e: Any exception raised during the process.
        """
        try:
            # Initialize two lists for final comparison file path and output final comparison sheet names
            paths = [self.out_final_cig, self.out_final_ecig, self.out_final_cgr, self.out_final_otp]
            sheets = self.out_summary_sheets[1:5]
            # Check if these two lists have the same length
            if len(paths) != len(sheets):
                raise ValueError("The number of fc file paths and output sheet names do not match.")
            # Loop through all the categories to create final comparison worksheets
            for input_fc, sheet_fc in zip(paths, sheets):
                self._create_fc_sheet(wb, input_fc, sheet_fc)
        except Exception as e:
            console.log(f"Error in _create_fc_sheets(): {e}")
            raise e
        else:
            console.log("All of the FC worksheets have been created successfully.")
        
    def _add_col_concatenated(self, wb: Workbook) -> pd.DataFrame:
        """Add the 'Concatenated' column to the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to get the data from.

        Returns:
            pd.DataFrame: the final dataframe with the 'Concatenated' column.
        """
        # Initialize an empty dataframe
        df_final = pd.DataFrame()
        # Get the data from worksheets cig, blu, cgr, otp as dataframes
        for sheet in self.out_summary_sheets[1:5]:
            ws = wb[sheet]
            # Get the values in column A starting from row 6
            data1 = [cell[0].value for cell in ws.iter_rows(min_row=6, min_col=1, max_col=1)]
            # Put the list into a dataframe with appropriate column name
            df = pd.DataFrame(data1, columns=[self.col_curr_db[0]])
            # Concatenate the distributor name with the sheet name
            df_final = pd.concat([df_final, df], ignore_index=True)
        # Remove all duplicates
        df_final.drop_duplicates(inplace=True)
        # Remove all thw null values
        df_final.dropna(inplace=True)
        # Sort the dataframe by the column 'Concatenated'
        df_final.sort_values(by=self.col_curr_db[0], inplace=True)
        # Return the final dataframe
        return df_final
    
    def _add_cols_dist_cust(self, df_final: pd.DataFrame) -> pd.DataFrame:
        """Add the 'Distributor Name' and 'Customer Number' columns to the final deliverable summary report.

        Args:
            df_final (pd.DataFrame): The final dataframe to add the columns to.

        Returns:
            pd.DataFrame: the final dataframe with the new columns.
        """
        # Split 'Concatenated' column while ensuring it only splits at the first "("
        split_df = df_final[self.col_curr_db[0]].str.split("(", n=1, expand=True)
        # Check if all rows result in two parts
        if split_df.shape[1] < 2:
            split_df[1] = pd.NA  # Ensure there are always two columns
        # Check each row for issues and print details if problematic
        problematic_rows = split_df[split_df[1].isna()]
        if not problematic_rows.empty:
            console.log("Rows that did not split correctly (missing or misplaced '('):")
            console.log(problematic_rows)
        # Assign cleaned-up columns back to the dataframe
        df_final[self.col_summary[0]] = split_df[0].str.strip()
        df_final[self.col_summary[1]] = split_df[1].str.replace(")", "", regex=False).str.strip() if split_df.shape[1] > 1 else ""
        return df_final
    
    def _add_cols_impacted(self, wb: Workbook, df_final: pd.DataFrame) -> pd.DataFrame:
        """Add the 'Impacted' columns to the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to get the data from.
            df_final (pd.DataFrame): The final dataframe to add the columns to.

        Returns:
            pd.DataFrame: the final dataframe with the new columns.
        """
        # Dictionary to store values from each sheet's column A for quick lookup
        sheet_values = {}
        # Pre-fetch and store column A values from each sheet
        for sheet in self.out_summary_sheets[1:5]:
            ws = wb[sheet]
            # Store all values from column A starting from row 6 in a set for quick lookup
            sheet_values[sheet] = {cell[0].value for cell in ws.iter_rows(min_row=6, min_col=1, max_col=1)}
        # Loop through each row in column 'Concatenated' in df_final
        for index, row in df_final.iterrows():
            # Get the value in 'Concatenated' column
            value = row[self.col_curr_db[0]]
            # Check this value against each sheet's column A values
            for sheet in sheet_values:
                # If value is found in the set of sheet values, assign 'X', otherwise assign ''
                df_final.loc[index, sheet] = "X" if value in sheet_values[sheet] else ""
        # Return the final dataframe
        return df_final
    
    def _add_cols_rsd_vol(self, wb: Workbook, df: pd.DataFrame) -> pd.DataFrame:
        """Add the RSD Volume columns 'Additions' and 'Decreases' to the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to get the data from.
            df (pd.DataFrame): The final dataframe to add the columns to.

        Raises:
            e: Any exception raised during the process.

        Returns:
            pd.DataFrame: the final dataframe with the new columns.
        """
        try:
            # Adding new columns to the DataFrame
            df['Additions'] = self.SPACE
            df['Decreases'] = self.SPACE
            # Loop through each row in DataFrame
            for index, row in df.iterrows():
                additions_flag = self.SPACE
                decreases_flag = self.SPACE
                # Check each worksheet
                for sheet_name in self.out_summary_sheets[1:5]:
                    ws = wb[sheet_name]
                    # Get all values in column A and corresponding values in column 19 from row 6 onwards
                    col1_values = [cell[0].value for cell in ws.iter_rows(min_row=6, min_col=1, max_col=1)]
                    col19_values = [cell[0].value for cell in ws.iter_rows(min_row=6, min_col=19, max_col=19)]
                    # Check conditions and update flags
                    for i, value in enumerate(col1_values):
                        if value == row[self.col_curr_db[0]]:
                            if col19_values[i] is not None:
                                if col19_values[i] > 0:
                                    additions_flag = 'X'
                                elif col19_values[i] < 0:
                                    decreases_flag = 'X'
                # Set flags in df
                df.at[index, 'Additions'] = additions_flag
                df.at[index, 'Decreases'] = decreases_flag
            return df
        except Exception as e:
            console.log(f"Error in _add_cols_rsd_vol(): {e}")
            raise e
        
    def _add_cols_vol_impact(self, wb: Workbook, df: pd.DataFrame) -> pd.DataFrame:
        """Add the 'Volume Impact' columns to the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to get the data from.
            df (pd.DataFrame): The final dataframe to add the columns to.

        Returns:
            pd.DataFrame: the final dataframe with the new columns.
        """
        sheets = {
            self.out_summary_sheets[1]: (self.col_summary[10], 4),
            self.out_summary_sheets[2]: (self.col_summary[11], 4),
            self.out_summary_sheets[3]: (self.col_summary[12], 4),
            self.out_summary_sheets[4]: (self.col_summary[13], 4)
        }
        # Initialize new columns with default blank space
        for sheet, info in sheets.items():
            df[info[0]] = self.SPACE  # info[0] is the column name

        # Loop through each row in the DataFrame
        for index, row in df.iterrows():
            value = row[self.col_curr_db[0]]
            # Check each sheet specified
            for sheet, info in sheets.items():
                ws = wb[sheet]
                column_name, column_index = info
                found = False
                # Scan through column A in the worksheet
                for cell in ws['A']:  # Assuming data starts from row 1, adjust range as necessary
                    if cell.value == value:
                        # Exact match found, now get value from the specified column
                        target_value = ws.cell(row=cell.row, column=column_index).value
                        if target_value is not None:
                            df.at[index, column_name] = target_value
                            found = True
                            break
                # If no match was found, ensure the column is set to a blank space
                if not found:
                    df.at[index, column_name] = self.SPACE
        return df
                
    def _add_cols_week_change(self, wb: Workbook, df: pd.DataFrame) -> pd.DataFrame:
        """Add the 'Week Change' columns to the final dataframe.

        Args:
            wb (Workbook): The workbook to get the data from.
            df (pd.DataFrame): The final dataframe to add the columns to.

        Returns:
            pd.DataFrame: the final dataframe with the new columns.
        """
        # Define the sheets and their corresponding columns for 'wk' and 'date'
        sheet_info = {
            self.out_summary_sheets[1]: {'wk_col': 21, 'date_col': 24},
            self.out_summary_sheets[2]: {'wk_col': 21, 'date_col': 24},
            self.out_summary_sheets[3]: {'wk_col': 21, 'date_col': 24},
            self.out_summary_sheets[4]: {'wk_col': 21, 'date_col': 24}
        }
        # Initialize new columns in the DataFrame
        for category, cols in sheet_info.items():
            df[f'{category} wk'] = None  # Initialize week columns
            df[f'{category} date'] = None  # Initialize date columns
        # Loop through each row in the DataFrame
        for index, row in df.iterrows():
            value = row[self.col_curr_db[0]]
            # Check each sheet
            for category, info in sheet_info.items():
                ws = wb[category]
                wk_col = info['wk_col']
                date_col = info['date_col']
                # Scan through column A in the worksheet
                for cell in ws['A']:
                    if cell.value == value:
                        # Match found, retrieve week and date
                        wk_cell = ws.cell(row=cell.row, column=wk_col)
                        date_cell = ws.cell(row=cell.row, column=date_col)
                        df.at[index, f'{category} wk'] = wk_cell.value
                        df.at[index, f'{category} date'] = date_cell.value
                        # Stop looking once a match is found for this category
                        break
        # Return the final DataFrame
        return df
    
    def _add_col_weeks_occurred(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add the 'Weeks Occurred' column to the final deliverable summary report.

        Args:
            df (pd.DataFrame): The final dataframe to add the column to.

        Returns:
            pd.DataFrame: the final dataframe with the new column.
        """
        df['Weeks Occurred'] = np.where(df['cig wk'].notna(), df['cig date'], np.where(df['blu wk'].notna(), df['blu date'], np.where(df['cgr wk'].notna(), df['cgr date'], df['otp date'])))
        #df['Weeks Occurred'] = np.where(df['cig wk'].notna(), df['cig date'], np.where(df['blu wk'].notna(), df['blu date'], df['cgr date']))
        return df
    
    def _add_cols_wdc_changes(self, wb: Workbook, df: pd.DataFrame) -> pd.DataFrame:
        """Add the '*wdc' columns to the final dataframe.

        Args:
            wb (Workbook): The workbook to get the data from.
            df (pd.DataFrame): The final dataframe to add the columns to.

        Returns:
            pd.DataFrame: the final dataframe with the new columns.
        """
        # Define the mapping of columns and sheets
        categories = {
            self.out_summary_sheets[5]: (self.col_summary[1], 'cig wk', self.out_summary_sheets[5]),
            self.out_summary_sheets[6]: (self.col_summary[1], 'blu wk', self.out_summary_sheets[6]),
            self.out_summary_sheets[7]: (self.col_summary[1], 'cgr wk', self.out_summary_sheets[7]),
            self.out_summary_sheets[8]: (self.col_summary[1], 'otp wk', self.out_summary_sheets[8])
        }
        # Process each category
        for key, (cust_num_col, week_col, sheet_name) in categories.items():
            # Initialize the new column with None
            df[key] = None
            # Get the corresponding wdc worksheet
            ws = wb[sheet_name]
            # Create a dictionary for fast lookup
            lookup_dict = {(str(row[2]) + str(row[5])): row[3] for row in ws.iter_rows(min_row=2, values_only=True)}
            # Perform the concatenation and lookup
            for index, row in df.iterrows():
                if pd.notna(row[week_col]):
                    lookup_value = f"{row[cust_num_col]}{row[week_col]}"
                    df.at[index, key] = lookup_dict.get(lookup_value, None)
        return df
        
    def _add_col_reason(self, wb: Workbook, df: pd.DataFrame) -> pd.DataFrame:
        """Add the 'Reason' column to the final deliverable summary report.

        Args:
            df (pd.DataFrame): The final dataframe to add the column to.

        Raises:
            ValueError: _description_

        Returns:
            pd.DataFrame: the final dataframe with the new column.
        """
        # Reset the index of the DataFrame
        df = df.reset_index(drop=True)
        # df[self.col_summary[1]] = df[self.col_summary[1]].astype(np.int64)

        # Get all reason data into single dataframe
        reason_sheet_names = [self.out_summary_sheets[5], self.out_summary_sheets[6], self.out_summary_sheets[7], self.out_summary_sheets[8]]
        reason_data_frames = []
        for sheet_name in reason_sheet_names:
            # Get individual sheet data
            sheet_data = pd.DataFrame(wb[sheet_name].values)
            sheet_data.columns = sheet_data.iloc[0]
            sheet_data = sheet_data[1:].reset_index(drop=True)
            reason_data_frames.append(sheet_data)
        # Concatenate all dataframe into one
        reason_data = pd.concat(reason_data_frames, ignore_index=True)
        reason_data = reason_data.reset_index(drop=True)

        # Search for reason values in the reason data and assign to the final dataframe
        reason_customer_num = "CUSTOMER_NUMBER"
        df[self.col_summary[9]] = ""
        for row in range(len(df.index)):
            customer_num = df.iloc[row, 2]
            reason_query = reason_data[reason_data[reason_customer_num] == customer_num]
            if len(reason_query.index) > 0:
                reason = reason_query.iloc[0, 3]
                if reason is None or reason == "" or pd.isna(reason):
                    continue
                df.at[row, self.col_summary[9]] = reason
            
        # Return the final DataFrame
        return df
        

    
        
    def _put_summary_data(self, wb:Workbook, ws: Worksheet):
        """Create the summary data and put it into the worksheet 'Summary'.

        Args:
            wb (Workbook): The workbook to get the data from.
            ws (Worksheet): The worksheet to put the data into.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Create a new column 'Concatenated' to store distributor name and customer number
            df_final = self._add_col_concatenated(wb)
            # Create columns 'Distributor Name' and 'Customer Number' by splitting the 'Concatenated' column
            df_final = self._add_cols_dist_cust(df_final)
            # Create columns 'Impacted Category' for each category
            df_final = self._add_cols_impacted(wb, df_final)
            # Create columns 'Additions' and 'Decreases'
            df_final = self._add_cols_rsd_vol(wb, df_final)
            # Create columns under 'Volume Impact'
            df_final = self._add_cols_vol_impact(wb, df_final)
            # Create columns under 'Week Change'
            df_final = self._add_cols_week_change(wb, df_final)
            # Create a new column 'Weeks Occurred'
            df_final = self._add_col_weeks_occurred(df_final)
            # Create columns under 'WDC changes info'
            df_final = self._add_cols_wdc_changes(wb, df_final)
            # Create a new column 'Reason'
            df_final = self._add_col_reason(wb, df_final)
            # Reorder the columns that we want to display in the final deliverable
            df_final = df_final[['Distributor Name', 'Customer Number', 'cig', 'blu', 'cgr', 'otp', 'Additions', 'Decreases', "Weeks Occurred", "Reason", 
            'Cigarettes (Ctns)', 'e-Cigs (Units)', 'Cigars (Sticks)', 'Otp (Sticks)']]
            # Put the data into the worksheet starting from cell A3
            TOB_ITG_Summary_Rpt.dataframe_to_excel(df_final, ws, skip_rows=2)
        except Exception as e:
            console.log(f"Error in _put_summary_data(): {e}")
            raise e
        
        
    def _set_style_summary(self, ws: Worksheet):
        """Set the style of the worksheet Summary_Rpt.

        Args:
            ws (Worksheet): The worksheet to set the style for.
        """
        # Set consistent style for all columns starting from row 3
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Set consistent style for column 1
                if cell.column == 1:
                    TOB_ITG_Summary_Rpt.set_cell_style(cell, None, name=self.calibri, size=11, horizontal=None, border=self.thin_border)
                # Set consistent style for the rest of the columns
                else:
                    TOB_ITG_Summary_Rpt.set_cell_style(cell, None, name=self.calibri, size=11, border=self.thin_border)
                # Set the number format for column H to show only the date
                if cell.column == 8:
                    cell.number_format = 'MM/DD/YYYY'
        # Automatically adjust the column width of the worksheet
        TOB_ITG_Summary_Rpt.auto_adjust_column_width(ws, 2)
        
                               
    def _create_summary_sheet(self, wb: Workbook):
        """Create the summary sheet in the final deliverable summary report.

        Args:
            wb (Workbook): The workbook to create the worksheet in.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Load the template for the summary sheet
            ws = wb[self.out_summary_sheets[0]]
            # Put data into the worksheet
            self._put_summary_data(wb, ws)
            # Set the style of the worksheet
            self._set_style_summary(ws)         
        except Exception as e:
            console.log(f"Error in _create_summary_sheet(): {e}")
            raise e
        else:
            console.log(f"Worksheet '{self.out_summary_sheets[0]}' has been created successfully.")
 
    def create_final_summary(self, template: str, output: str):
        """Create the final deliverable summary report.

        Args:
            template (str): The template file path for the final deliverable.
            output (str): The output file path for the final deliverable.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Load the workbook for the final deliverable template file
            wb = load_workbook(template)
            # Create the wdc worksheets for all categories
            self._create_wdc_sheets(wb)
            # Create the Final_Comparison sheets for all categories
            self._create_fc_sheets(wb)
            # Create the summary sheet
            self._create_summary_sheet(wb)
            # Save and close the workbook
            TOB_ITG_Summary_Rpt.close_wb(wb, output)
        except Exception as e:
            console.log(f"Error in create_final_summary(): {e}")
            raise e
        else:
            console.log(f"Final deliverable '{output}' has been created successfully.")
            
    
    def run(self):
        """Run the entire process to create volume changes report, WDC report, final deliverables for all category.

        Raises:
            e: Any exception raised during the process.
        """
        try:
            # Create volume changes dynamic report, wdc dynamic report, and final deliverable for cig, ecig, cgr, otp
            self.cig_job()
            self.ecig_job()
            self.cgr_job()
            self.otp_job()
            # Create Volume Changes Summary report
            self.create_final_summary(self.temp_final_summary, self.out_final_summary)
        except Exception as e:
            console.log(f"Error in run(): {e}\n{traceback.format_exc()}")
            raise e
